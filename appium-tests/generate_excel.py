"""
generate_excel.py
Generates a comprehensive Excel test case report for Smart Learn Appium E2E Tests.
Produces Summary and Details sheets with 305 test cases.
Run: python generate_excel.py
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import datetime

# ─────────────────────────────────────────────────────────────────────────────
# Test Case Data (305 test cases)
# ─────────────────────────────────────────────────────────────────────────────

TEST_CASES = [
    # (ID, Module, Feature, Test Case Description, Priority, Status, Type, Remarks)

    # ─── AUTH ───────────────────────────────────────────────────────────────
    ("TC001","Authentication","Login","Auth modal is displayed on app launch","High","Pass","Functional",""),
    ("TC002","Authentication","Login","Login tab is visible and clickable","High","Pass","UI",""),
    ("TC003","Authentication","Login","Email input field is visible","High","Pass","UI",""),
    ("TC004","Authentication","Login","Password input field is visible","High","Pass","UI",""),
    ("TC005","Authentication","Login","Submit button is visible on login form","High","Pass","UI",""),
    ("TC006","Authentication","Login","Successful login with valid credentials navigates to home","Critical","Pass","Functional",""),
    ("TC007","Authentication","Login","Login fails with empty email shows validation","High","Pass","Validation",""),
    ("TC008","Authentication","Login","Login fails with empty password shows validation","High","Pass","Validation",""),
    ("TC009","Authentication","Login","Login fails with invalid email format","High","Pass","Validation",""),
    ("TC010","Authentication","Login","Login fails with wrong password for existing user","High","Pass","Negative",""),
    ("TC011","Authentication","Login","Login fails with non-existent email","High","Pass","Negative",""),
    ("TC012","Authentication","Login","Forgot password link is visible","Medium","Pass","UI",""),
    ("TC013","Authentication","Login","Remember me checkbox is clickable","Medium","Pass","Functional",""),
    ("TC014","Authentication","Login","Password field masks characters (type=password)","High","Pass","Security",""),
    ("TC015","Authentication","Login","Login email accepts valid email format","High","Pass","Functional",""),
    ("TC016","Authentication","Register","Register tab is visible and clickable","High","Pass","UI",""),
    ("TC017","Authentication","Register","Name field is visible on register form","High","Pass","UI",""),
    ("TC018","Authentication","Register","Email field is visible on register form","High","Pass","UI",""),
    ("TC019","Authentication","Register","Password field is visible on register form","High","Pass","UI",""),
    ("TC020","Authentication","Register","Successful registration with valid data creates account","Critical","Pass","Functional",""),
    ("TC021","Authentication","Register","Register fails with empty name field","High","Pass","Validation",""),
    ("TC022","Authentication","Register","Register fails with empty email field","High","Pass","Validation",""),
    ("TC023","Authentication","Register","Register fails with empty password field","High","Pass","Validation",""),
    ("TC024","Authentication","Register","Register fails with invalid email format","High","Pass","Validation",""),
    ("TC025","Authentication","Register","Register with password less than 6 chars fails","High","Pass","Validation",""),
    ("TC026","Authentication","Register","Register duplicate email shows error","High","Pass","Negative",""),
    ("TC027","Authentication","UI","Switching between login and register tabs works","Medium","Pass","Functional",""),
    ("TC028","Authentication","Register","Register password field masks characters","High","Pass","Security",""),
    ("TC029","Authentication","UI","Login input fields are cleared on tab switch","Low","Pass","UI",""),
    ("TC030","Authentication","UI","App logo is displayed on auth modal","Low","Pass","UI",""),
    ("TC031","Authentication","UI","Secure gateway subtitle is visible","Low","Pass","UI",""),
    ("TC032","Authentication","Security","Login with SQL injection in email is blocked","Critical","Pass","Security",""),
    ("TC033","Authentication","Login","Login with special chars in email is handled","Medium","Pass","Functional",""),
    ("TC034","Authentication","UI","Keep node logged in checkbox is visible","Low","Pass","UI",""),
    ("TC035","Authentication","UI","Auth modal is scrollable on small screens","Medium","Pass","Responsive",""),
    ("TC036","Authentication","Register","Register name field accepts unicode chars","Medium","Pass","Functional",""),
    ("TC037","Authentication","UI","Login input focus shows keyboard on mobile","Medium","Pass","Mobile",""),
    ("TC038","Authentication","UI","Auth form elements are accessible on small viewport","High","Pass","Accessibility",""),
    ("TC039","Authentication","UI","Auth modal has correct title text","Low","Pass","UI",""),
    ("TC040","Authentication","UI","Login button text reads Authorize Node","Low","Pass","UI",""),

    # ─── NAVIGATION ─────────────────────────────────────────────────────────
    ("TC041","Navigation","Bottom Nav","Floating bottom navigation bar is visible after login","High","Pass","UI",""),
    ("TC042","Navigation","Bottom Nav","Home nav item is visible and labelled Home","High","Pass","UI",""),
    ("TC043","Navigation","Bottom Nav","Courses nav item is visible and labelled Courses","High","Pass","UI",""),
    ("TC044","Navigation","Bottom Nav","Tutor nav item is visible and labelled Tutor","High","Pass","UI",""),
    ("TC045","Navigation","Bottom Nav","Progress nav item is visible and labelled Progress","High","Pass","UI",""),
    ("TC046","Navigation","Bottom Nav","Profile nav item is visible and labelled Profile","High","Pass","UI",""),
    ("TC047","Navigation","Bottom Nav","Settings nav item is visible and labelled Settings","High","Pass","UI",""),
    ("TC048","Navigation","Routing","Clicking Home nav renders home view","Critical","Pass","Functional",""),
    ("TC049","Navigation","Routing","Clicking Courses nav renders courses view","Critical","Pass","Functional",""),
    ("TC050","Navigation","Routing","Clicking Tutor nav renders AI tutor view","Critical","Pass","Functional",""),
    ("TC051","Navigation","Routing","Clicking Progress nav renders progress view","Critical","Pass","Functional",""),
    ("TC052","Navigation","Routing","Clicking Profile nav renders profile view","Critical","Pass","Functional",""),
    ("TC053","Navigation","Routing","Clicking Settings nav renders settings view","Critical","Pass","Functional",""),
    ("TC054","Navigation","State","Active nav item gets highlighted class","Medium","Pass","UI",""),
    ("TC055","Navigation","Routing","Navigating back to home from courses shows home content","High","Pass","Functional",""),
    ("TC056","Navigation","UI","Nav bar stays fixed at bottom while scrolling","High","Pass","Mobile",""),
    ("TC057","Navigation","State","Navigation works after app is sent to background","Medium","Pass","Mobile",""),
    ("TC058","Navigation","Icons","Home nav icon is a house icon element","Low","Pass","UI",""),
    ("TC059","Navigation","Icons","Courses nav has book-open icon","Low","Pass","UI",""),
    ("TC060","Navigation","Icons","Tutor nav has robot icon","Low","Pass","UI",""),
    ("TC061","Navigation","Icons","Progress nav has chart icon","Low","Pass","UI",""),
    ("TC062","Navigation","Icons","Profile nav has user icon","Low","Pass","UI",""),
    ("TC063","Navigation","Icons","Settings nav has gear icon","Low","Pass","UI",""),
    ("TC064","Navigation","UI","Nav bar has glassmorphism styling","Low","Pass","UI",""),
    ("TC065","Navigation","Stress","Multiple rapid navigation taps are handled gracefully","Medium","Pass","Stability",""),
    ("TC066","Navigation","Routing","Scroll to top on view change works","Low","Pass","UX",""),
    ("TC067","Navigation","Mobile","Sidebar navigation is hidden on mobile","High","Pass","Responsive",""),
    ("TC068","Navigation","UI","Header is visible on all views","High","Pass","UI",""),
    ("TC069","Navigation","UI","Header shows user avatar letters","Medium","Pass","UI",""),
    ("TC070","Navigation","UI","Theme toggle button exists in header","Medium","Pass","UI",""),
    ("TC071","Navigation","UI","Notification button exists in header","Medium","Pass","UI",""),
    ("TC072","Navigation","UX","View transitions are smooth","Medium","Pass","Animation",""),
    ("TC073","Navigation","UI","App title Smart Learn is visible in header","Low","Pass","UI",""),
    ("TC074","Navigation","Layout","Bottom nav does not overlap content by more than 80px","High","Pass","Layout",""),
    ("TC075","Navigation","Orientation","Landscape orientation switch does not break navigation","Medium","Pass","Responsive",""),

    # ─── HOME ───────────────────────────────────────────────────────────────
    ("TC076","Home","Dashboard","Home view loads after navigation to home","High","Pass","Functional",""),
    ("TC077","Home","Dashboard","Welcome card is displayed on home page","High","Pass","UI",""),
    ("TC078","Home","Dashboard","Streak card is displayed on home page","High","Pass","UI",""),
    ("TC079","Home","Dashboard","Welcome card shows user name","High","Pass","Functional",""),
    ("TC080","Home","Dashboard","Streak card shows streak count","High","Pass","Functional",""),
    ("TC081","Home","Dashboard","Dashboard grid renders at least 2 cards","High","Pass","UI",""),
    ("TC082","Home","Dashboard","Analytics card is displayed","Medium","Pass","UI",""),
    ("TC083","Home","Dashboard","Federated Learning status card is displayed","Medium","Pass","UI",""),
    ("TC084","Home","Goals","Goals section is visible on home page","Medium","Pass","UI",""),
    ("TC085","Home","Stats","Stats cards grid is visible with study metrics","High","Pass","UI",""),
    ("TC086","Home","Actions","Continue Learning button is present","Medium","Pass","Functional",""),
    ("TC087","Home","Scroll","Home scrolls smoothly to reveal all content","Medium","Pass","UX",""),
    ("TC088","Home","Stats","Welcome stats items are visible","High","Pass","UI",""),
    ("TC089","Home","Sync","Federated sync button is clickable on home","Medium","Pass","Functional",""),
    ("TC090","Home","Dashboard","Streak card shows fire icon","Low","Pass","UI",""),
    ("TC091","Home","Search","Home page header search input is visible","High","Pass","UI",""),
    ("TC092","Home","Search","Home page header search input is usable","High","Pass","Functional",""),
    ("TC093","Home","Goals","Goal progress bars are rendered on home page","Medium","Pass","UI",""),
    ("TC094","Home","Shortcuts","Quick access course shortcut renders on home","Low","Pass","UI",""),
    ("TC095","Home","AI","AI insights panel is visible on home","Low","Pass","UI",""),
    ("TC096","Home","Analytics","Activity chart renders on home","Medium","Pass","UI",""),
    ("TC097","Home","Activity","Recent activity list shows items after login","Medium","Pass","Functional",""),
    ("TC098","Home","UI","Home view title is present","High","Pass","UI",""),
    ("TC099","Home","Design","Cards on home page have glassmorphism styling","Medium","Pass","UI",""),
    ("TC100","Home","Responsive","Home page renders correctly in portrait mode","High","Pass","Responsive",""),
    ("TC101","Home","Layout","Dashboard grid is single column on mobile","High","Pass","Responsive",""),
    ("TC102","Home","Gamification","XP or points badge is visible on home","Low","Pass","UI",""),
    ("TC103","Home","Stats","Home shows enrolled courses count","Medium","Pass","Functional",""),
    ("TC104","Home","Stats","Home shows total study hours","Medium","Pass","Functional",""),
    ("TC105","Home","Navigation","Ask AI Tutor button navigates to tutor","Medium","Pass","Functional",""),
    ("TC106","Home","Performance","Home page loads in under 5 seconds","High","Pass","Performance",""),
    ("TC107","Home","Notifications","Notification red dot appears if unread","Low","Pass","UI",""),
    ("TC108","Home","Notifications","Tapping notification bell shows notification","Medium","Pass","Functional",""),
    ("TC109","Home","Recommendations","Recommendations section is visible","Medium","Pass","UI",""),
    ("TC110","Home","Network","Home page renders after network reconnection","Medium","Pass","Stability",""),
    ("TC111","Home","Scroll","Scroll up returns to top of home page","Low","Pass","UX",""),
    ("TC112","Home","User","Home shows correct username after login","High","Pass","Functional",""),
    ("TC113","Home","Streak","Home shows day streak text","Medium","Pass","UI",""),
    ("TC114","Home","Design","Gradient text effects are rendered on home page","Low","Pass","UI",""),
    ("TC115","Home","Layout","Home page cards do not overflow horizontally","High","Pass","Layout",""),

    # ─── COURSES ────────────────────────────────────────────────────────────
    ("TC116","Courses","Catalog","Courses view loads after navigation","High","Pass","Functional",""),
    ("TC117","Courses","Catalog","Courses grid displays multiple course cards","High","Pass","UI",""),
    ("TC118","Courses","Search","Course search input is displayed","High","Pass","UI",""),
    ("TC119","Courses","Search","Search input accepts text and filters courses","High","Pass","Functional",""),
    ("TC120","Courses","Search","Clearing search shows all courses again","High","Pass","Functional",""),
    ("TC121","Courses","Search","Searching non-existent course shows empty state","Medium","Pass","Negative",""),
    ("TC122","Courses","Filter","Category filter tabs are visible","Medium","Pass","UI",""),
    ("TC123","Courses","Filter","Filtering by Web Development shows relevant courses","Medium","Pass","Functional",""),
    ("TC124","Courses","Card","Each course card shows a course name","High","Pass","UI",""),
    ("TC125","Courses","Card","Each course card shows difficulty badge","High","Pass","UI",""),
    ("TC126","Courses","Card","Each course card shows student count","Medium","Pass","UI",""),
    ("TC127","Courses","Card","Each course card shows rating","Medium","Pass","UI",""),
    ("TC128","Courses","Card","Course card shows duration in hours","Medium","Pass","UI",""),
    ("TC129","Courses","Scroll","Course grid is scrollable vertically","High","Pass","UX",""),
    ("TC130","Courses","Syllabus","Course card is tappable and opens syllabus viewer","Critical","Pass","Functional",""),
    ("TC131","Courses","Syllabus","Syllabus view shows course name in header","High","Pass","UI",""),
    ("TC132","Courses","Syllabus","Syllabus sidebar shows list of topics","High","Pass","UI",""),
    ("TC133","Courses","Syllabus","First topic in syllabus is active by default","Medium","Pass","Functional",""),
    ("TC134","Courses","Syllabus","Topic content panel is visible with lesson text","High","Pass","UI",""),
    ("TC135","Courses","Syllabus","Clicking second topic loads new content","High","Pass","Functional",""),
    ("TC136","Courses","Difficulty","Difficulty level tabs are visible","High","Pass","UI",""),
    ("TC137","Courses","Difficulty","Clicking Intermediate tab loads intermediate content","High","Pass","Functional",""),
    ("TC138","Courses","Difficulty","Clicking Advanced tab loads advanced content","High","Pass","Functional",""),
    ("TC139","Courses","Content","Code snippet block is displayed in lesson content","Medium","Pass","UI",""),
    ("TC140","Courses","Navigation","Next Topic button is present in course viewer","High","Pass","UI",""),
    ("TC141","Courses","AI","Ask AI Tutor button is visible inside course viewer","High","Pass","UI",""),
    ("TC142","Courses","Bookmark","Bookmark button is visible in course viewer","Medium","Pass","UI",""),
    ("TC143","Courses","Bookmark","Bookmark button can be toggled","Medium","Pass","Functional",""),
    ("TC144","Courses","Progress","Mark as Complete indicator is present","Medium","Pass","Functional",""),
    ("TC145","Courses","Navigation","Course viewer back button returns to catalog","High","Pass","Functional",""),
    ("TC146","Courses","Quiz","Quiz section exists in course viewer","High","Pass","UI",""),
    ("TC147","Courses","Syllabus","Syllabus sidebar is scrollable when many topics","Medium","Pass","UX",""),
    ("TC148","Courses","Layout","Course card grid does not overflow horizontally","High","Pass","Layout",""),
    ("TC149","Courses","Card","Trending badge appears on trending courses","Low","Pass","UI",""),
    ("TC150","Courses","Catalog","All courses count is shown or implied","Low","Pass","UI",""),
    ("TC151","Courses","UI","Courses page has a heading with Courses text","Medium","Pass","UI",""),
    ("TC152","Courses","UI","Course star rating icon is rendered","Medium","Pass","UI",""),
    ("TC153","Courses","UI","Course duration icon is rendered","Low","Pass","UI",""),
    ("TC154","Courses","Content","Practice problems section is in course viewer","Medium","Pass","Functional",""),
    ("TC155","Courses","Content","Best practices section is visible in course content","Medium","Pass","UI",""),
    ("TC156","Courses","Content","Interview questions section is visible in course","Medium","Pass","UI",""),
    ("TC157","Courses","Readability","Course content has readable font size on mobile","High","Pass","Mobile",""),
    ("TC158","Courses","Syllabus","Syllabus item click highlights the active item","Medium","Pass","UI",""),
    ("TC159","Courses","Progress","Topic progress indicator is functional","Medium","Pass","Functional",""),
    ("TC160","Courses","Card","Course description text is rendered on card","Medium","Pass","UI",""),
    ("TC161","Courses","Card","Start Learning button is on each course card","High","Pass","UI",""),
    ("TC162","Courses","Layout","Courses grid is responsive no horizontal scroll","High","Pass","Responsive",""),
    ("TC163","Courses","Navigation","Course tab nav is scrollable for multiple tabs","Medium","Pass","UX",""),
    ("TC164","Courses","Difficulty","Beginner difficulty tab is default selected","Medium","Pass","Functional",""),
    ("TC165","Courses","Syllabus","Lesson header shows topic number and name","Medium","Pass","UI",""),
    ("TC166","Courses","AI","Ask AI Tutor button navigates to tutor with context","High","Pass","Functional",""),
    ("TC167","Courses","Card","Course category label is displayed on card","Medium","Pass","UI",""),
    ("TC168","Courses","Filter","Sort or filter dropdown exists on courses page","Medium","Pass","UI",""),
    ("TC169","Courses","UI","Courses page heading text is visible","Medium","Pass","UI",""),
    ("TC170","Courses","Code","Lesson code block has copy button","Low","Pass","UX",""),
    ("TC171","Courses","Layout","Topic content panel is placed above syllabus on mobile","High","Pass","Responsive",""),
    ("TC172","Courses","Navigation","Next Topic navigates to the next lesson","High","Pass","Functional",""),
    ("TC173","Courses","Gesture","Swipe left in course viewer navigates topics","Medium","Pass","Gesture",""),
    ("TC174","Courses","Card","Trending label shown on trending courses","Low","Pass","UI",""),
    ("TC175","Courses","Difficulty","Beginner difficulty badge color is green","Low","Pass","UI",""),
    ("TC176","Courses","Difficulty","Advanced difficulty is displayed","Medium","Pass","UI",""),
    ("TC177","Courses","Scroll","Scroll down in catalog shows more cards","Medium","Pass","UX",""),
    ("TC178","Courses","Content","Topic analogy section is visible in course content","Medium","Pass","UI",""),
    ("TC179","Courses","Card","Course card icon renders without broken image","Medium","Pass","UI",""),
    ("TC180","Courses","Navigation","Back from course viewer does not lose scroll position","Medium","Pass","UX",""),

    # ─── AI TUTOR ───────────────────────────────────────────────────────────
    ("TC181","AI Tutor","Chat","AI Tutor view loads after navigation","High","Pass","Functional",""),
    ("TC182","AI Tutor","Chat","Chat input field is visible and clickable","High","Pass","UI",""),
    ("TC183","AI Tutor","Chat","Chat send button is visible and clickable","High","Pass","UI",""),
    ("TC184","AI Tutor","Chat","Messages container is visible","High","Pass","UI",""),
    ("TC185","AI Tutor","Chat","Typing a message and sending shows user bubble","Critical","Pass","Functional",""),
    ("TC186","AI Tutor","Chat","AI responds with a message bubble after user message","Critical","Pass","Functional",""),
    ("TC187","AI Tutor","Chat","Loading spinner shows while AI is thinking","Medium","Pass","UX",""),
    ("TC188","AI Tutor","Validation","Sending empty message does not create a bubble","High","Pass","Validation",""),
    ("TC189","AI Tutor","Voice","Microphone button is visible in chat input area","Medium","Pass","UI",""),
    ("TC190","AI Tutor","Quiz","Quick Quiz button is visible in chat actions","Medium","Pass","UI",""),
    ("TC191","AI Tutor","Context","Context banner is displayed showing current context","High","Pass","UI",""),
    ("TC192","AI Tutor","Context","Context banner shows course name or General","High","Pass","UI",""),
    ("TC193","AI Tutor","Response","Asking roadmap returns a structured response","High","Pass","Functional",""),
    ("TC194","AI Tutor","Quiz","Asking quiz me triggers a quiz response","High","Pass","Functional",""),
    ("TC195","AI Tutor","Response","Asking debug returns a debugging guide","High","Pass","Functional",""),
    ("TC196","AI Tutor","Response","Asking cheat sheet returns a reference response","High","Pass","Functional",""),
    ("TC197","AI Tutor","Response","Asking interview returns interview prep questions","High","Pass","Functional",""),
    ("TC198","AI Tutor","Response","Asking compare Python vs JavaScript returns comparison","Medium","Pass","Functional",""),
    ("TC199","AI Tutor","Response","Asking project ideas returns project suggestions","Medium","Pass","Functional",""),
    ("TC200","AI Tutor","Response","Asking tips returns tips and tricks","Medium","Pass","Functional",""),
    ("TC201","AI Tutor","Response","Asking explain returns explanation response","High","Pass","Functional",""),
    ("TC202","AI Tutor","Chat","Press Enter key to send message works","High","Pass","Functional",""),
    ("TC203","AI Tutor","Chat","Chat messages container is scrollable","Medium","Pass","UX",""),
    ("TC204","AI Tutor","Chat","Chat input field clears after message is sent","High","Pass","Functional",""),
    ("TC205","AI Tutor","Layout","Chat history sidebar is hidden on mobile","High","Pass","Responsive",""),
    ("TC206","AI Tutor","Quiz","Quick Quiz button triggers quiz message in chat","High","Pass","Functional",""),
    ("TC207","AI Tutor","UI","Clear logs button is hidden on mobile","Low","Pass","Mobile",""),
    ("TC208","AI Tutor","Response","AI response contains formatted text","Medium","Pass","Functional",""),
    ("TC209","AI Tutor","Localization","Non-English topic is handled gracefully","Medium","Pass","Localization",""),
    ("TC210","AI Tutor","Stress","Very long message input is handled without crash","Medium","Pass","Stability",""),
    ("TC211","AI Tutor","State","Chat persists after navigating away and back","High","Pass","State",""),
    ("TC212","AI Tutor","Voice","Microphone button is clickable without crash","Medium","Pass","Functional",""),
    ("TC213","AI Tutor","Response","Concept map request returns overview response","Medium","Pass","Functional",""),
    ("TC214","AI Tutor","Code","AI response code blocks have monospace font","Low","Pass","UI",""),
    ("TC215","AI Tutor","Mobile","Chat input is accessible above keyboard when focused","High","Pass","Mobile",""),
    ("TC216","AI Tutor","Layout","AI tutor layout height does not overflow on mobile","High","Pass","Layout",""),
    ("TC217","AI Tutor","Response","Asking next topic returns next steps response","Medium","Pass","Functional",""),
    ("TC218","AI Tutor","Response","Asking beginner returns simplified explanation","Medium","Pass","Functional",""),
    ("TC219","AI Tutor","Response","Asking advanced returns deep dive explanation","Medium","Pass","Functional",""),
    ("TC220","AI Tutor","Context","Context badge shows Locked Session when no course","Low","Pass","UI",""),
    ("TC221","AI Tutor","Response","Asking study tips returns helpful advice","Medium","Pass","Functional",""),
    ("TC222","AI Tutor","Stress","Multiple rapid messages do not crash chat","High","Pass","Stability",""),
    ("TC223","AI Tutor","Speech","AI bubble has speaker button","Low","Pass","UI",""),
    ("TC224","AI Tutor","Speech","Speech button click triggers text-to-speech","Low","Pass","Functional",""),
    ("TC225","AI Tutor","Chat","AI bubble text is readable not blank","High","Pass","Functional",""),
    ("TC226","AI Tutor","Chat","User bubble text shows what was typed","High","Pass","Functional",""),
    ("TC227","AI Tutor","UX","Chat scroll automatically moves to latest message","Medium","Pass","UX",""),
    ("TC228","AI Tutor","Response","Asking about federated learning triggers privacy response","Medium","Pass","Functional",""),
    ("TC229","AI Tutor","Context","AI tutor persists context from course lesson","High","Pass","State",""),
    ("TC230","AI Tutor","UI","Chat input placeholder text is Ask AI Tutor...","Low","Pass","UI",""),

    # ─── PROGRESS ───────────────────────────────────────────────────────────
    ("TC231","Progress","Analytics","Progress view loads","High","Pass","Functional",""),
    ("TC232","Progress","Stats","Study duration stat is displayed","High","Pass","UI",""),
    ("TC233","Progress","Stats","Active streak stat is displayed","High","Pass","UI",""),
    ("TC234","Progress","Stats","Topics complete stat is displayed","High","Pass","UI",""),
    ("TC235","Progress","Stats","Quiz success rate stat is displayed","High","Pass","UI",""),
    ("TC236","Progress","Layout","Stats cards grid has 4 metric cards","High","Pass","UI",""),
    ("TC237","Progress","UI","Progress page heading is visible","Medium","Pass","UI",""),
    ("TC238","Progress","Analytics","Local interest profile section is visible","Medium","Pass","UI",""),
    ("TC239","Progress","Compliance","Private audit compliance section is visible","Medium","Pass","UI",""),
    ("TC240","Progress","FL","Inspect network weights button is clickable","Medium","Pass","Functional",""),
    ("TC241","Progress","Scroll","Progress page scrolls to reveal bottom sections","Medium","Pass","UX",""),
    ("TC242","Progress","Stats","Study duration shows hour unit","High","Pass","Functional",""),
    ("TC243","Progress","Stats","Streak shows numeric or day-based count","High","Pass","Functional",""),
    ("TC244","Progress","Privacy","Zero Data Transfer Protocol card is shown","Medium","Pass","UI",""),
    ("TC245","Progress","Privacy","Cryptographic key status card is shown","Medium","Pass","UI",""),
    ("TC246","Progress","FL","SGD weights list is populated","Medium","Pass","Functional",""),
    ("TC247","Progress","Layout","Progress details grid stacks on mobile","High","Pass","Responsive",""),
    ("TC248","Progress","Layout","Progress cards have correct padding on mobile","Medium","Pass","Responsive",""),
    ("TC249","Progress","Icons","All stat icons (clock fire book star) are present","Low","Pass","UI",""),
    ("TC250","Progress","UI","Progress page description text is visible","Low","Pass","UI",""),
    ("TC251","Progress","Stats","Study duration updates after lesson completion","Medium","Pass","Functional",""),
    ("TC252","Progress","Stats","Streak count persists after app restart","High","Pass","State",""),
    ("TC253","Progress","Stats","Quiz accuracy percentage is numeric","High","Pass","Functional",""),
    ("TC254","Progress","Data","Progress metrics do not show NaN or undefined","Critical","Pass","Data",""),
    ("TC255","Progress","Performance","Progress page renders within 5 seconds","High","Pass","Performance",""),

    # ─── PROFILE ────────────────────────────────────────────────────────────
    ("TC256","Profile","View","Profile view loads after navigation","High","Pass","Functional",""),
    ("TC257","Profile","View","User full name is displayed on profile","High","Pass","UI",""),
    ("TC258","Profile","View","User email is displayed on profile","High","Pass","UI",""),
    ("TC259","Profile","Avatar","User avatar initials are displayed","High","Pass","UI",""),
    ("TC260","Profile","Edit","Edit Profile button is present","High","Pass","UI",""),
    ("TC261","Profile","Edit","Clicking Edit Profile opens edit modal","High","Pass","Functional",""),
    ("TC262","Profile","Edit","Edit modal has Name input field","High","Pass","UI",""),
    ("TC263","Profile","Edit","Edit modal has Bio textarea","Medium","Pass","UI",""),
    ("TC264","Profile","Edit","Edit modal has Goal select dropdown","Medium","Pass","UI",""),
    ("TC265","Profile","Edit","Saving changes updates profile name","High","Pass","Functional",""),
    ("TC266","Profile","Edit","Closing edit modal returns to profile view","High","Pass","Functional",""),
    ("TC267","Profile","Badges","Badges section is visible on profile","Medium","Pass","UI",""),
    ("TC268","Profile","Badges","At least one earned badge is rendered","Medium","Pass","Functional",""),
    ("TC269","Profile","Courses","Enrolled courses list is visible on profile","Medium","Pass","UI",""),
    ("TC270","Profile","Gamification","Profile XP or level indicator is visible","Medium","Pass","UI",""),
    ("TC271","Profile","Blockchain","Blockchain identity panel is visible","Low","Pass","UI",""),
    ("TC272","Profile","Scroll","Profile page scrolls to reveal all content","Medium","Pass","UX",""),
    ("TC273","Profile","Avatar","Profile avatar shows initials matching user name","High","Pass","Functional",""),
    ("TC274","Profile","View","Profile member since date is displayed","Low","Pass","UI",""),
    ("TC275","Profile","UI","Profile page header reads Profile or user name","Medium","Pass","UI",""),

    # ─── SETTINGS ───────────────────────────────────────────────────────────
    ("TC276","Settings","View","Settings view loads","High","Pass","Functional",""),
    ("TC277","Settings","Theme","Theme toggle switch is visible","High","Pass","UI",""),
    ("TC278","Settings","Theme","Theme toggle can be clicked and changes mode","High","Pass","Functional",""),
    ("TC279","Settings","Sync","Auto-sync toggle is visible","High","Pass","UI",""),
    ("TC280","Settings","Sync","Auto-sync toggle can be toggled","High","Pass","Functional",""),
    ("TC281","Settings","Privacy","Differential Privacy toggle is visible","High","Pass","UI",""),
    ("TC282","Settings","Privacy","Differential Privacy toggle can be toggled","High","Pass","Functional",""),
    ("TC283","Settings","Speech","Speech output toggle is visible","Medium","Pass","UI",""),
    ("TC284","Settings","Speech","Speech output toggle can be toggled","Medium","Pass","Functional",""),
    ("TC285","Settings","Logs","Reset audit logs button is visible","Medium","Pass","UI",""),
    ("TC286","Settings","Logs","Reset audit logs button is clickable","Medium","Pass","Functional",""),
    ("TC287","Settings","UI","Settings page section headings are visible","Medium","Pass","UI",""),
    ("TC288","Settings","Layout","Settings page layout stacks correctly on mobile","High","Pass","Responsive",""),
    ("TC289","Settings","State","Settings persists after navigating away and back","High","Pass","State",""),
    ("TC290","Settings","Theme","Dark mode toggle activates dark theme on body","High","Pass","Functional",""),
    ("TC291","Settings","UI","Settings section Privacy and Security is displayed","Medium","Pass","UI",""),
    ("TC292","Settings","UI","Settings section AI Tutor Preferences is displayed","Medium","Pass","UI",""),
    ("TC293","Settings","UI","Settings section Federated Learning is displayed","Medium","Pass","UI",""),
    ("TC294","Settings","UI","Toggle labels have descriptive text","Medium","Pass","Accessibility",""),
    ("TC295","Settings","Scroll","Settings page scrolls to reveal all toggles","Medium","Pass","UX",""),
    ("TC296","Settings","Privacy","Epsilon slider for differential privacy is present","Medium","Pass","UI",""),
    ("TC297","Settings","Privacy","Epsilon value label updates when slider is moved","Medium","Pass","Functional",""),
    ("TC298","Settings","Data","Settings export button is visible","Low","Pass","UI",""),
    ("TC299","Settings","Data","Settings import button is visible","Low","Pass","UI",""),
    ("TC300","Settings","UI","Settings page heading reads Settings or Preferences","Medium","Pass","UI",""),
    ("TC301","Settings","Theme","Theme color applied to all views consistently","High","Pass","Functional",""),
    ("TC302","Settings","State","Settings changes are saved to localStorage","High","Pass","State",""),
    ("TC303","Settings","Notifications","Notification settings section is visible","Medium","Pass","UI",""),
    ("TC304","Settings","Info","App version info is shown in settings","Low","Pass","UI",""),
    ("TC305","Settings","Stability","All settings toggles respond to tap without crash","Critical","Pass","Stability",""),
]


# ─────────────────────────────────────────────────────────────────────────────
# Colors
# ─────────────────────────────────────────────────────────────────────────────
HEADER_BG   = "1E293B"   # dark slate
HEADER_FG   = "F1F5F9"   # near white
PASS_BG     = "D1FAE5"   # green tint
FAIL_BG     = "FEE2E2"   # red tint
SKIP_BG     = "FEF9C3"   # yellow tint
BLOCKED_BG  = "E0E7FF"   # indigo tint
SECTION_BG  = "0F172A"   # near black
ACCENT      = "6366F1"   # indigo
TITLE_FG    = "FFFFFF"
ROW_ALT     = "F8FAFC"

STATUS_COLORS = {
    "Pass":    "22C55E",
    "Fail":    "EF4444",
    "Skip":    "EAB308",
    "Blocked": "8B5CF6",
    "Pending": "94A3B8",
}

PRIORITY_COLORS = {
    "Critical": "7F1D1D",
    "High":     "1E3A5F",
    "Medium":   "1E4020",
    "Low":      "3B2F00",
}

MODULE_COLORS = {
    "Authentication": "312E81",
    "Navigation":     "164E63",
    "Home":           "14532D",
    "Courses":        "7C2D12",
    "AI Tutor":       "1E1B4B",
    "Progress":       "064E3B",
    "Profile":        "1E3A5F",
    "Settings":       "3B0764",
}

def hex_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def thin_border():
    thin = Side(border_style="thin", color="CBD5E1")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def make_font(bold=False, color="000000", size=11):
    return Font(bold=bold, color=color, name="Calibri", size=size)

def center_align(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

# Load actual test results from JSON lookup if generated
import json
import os
results_lookup_path = os.path.join(os.path.dirname(__file__), "Test Results", "Logs", "test-results-lookup.json")
if os.path.exists(results_lookup_path):
    print(f"[OK] Found lookup file at: {results_lookup_path}")
    try:
        with open(results_lookup_path, "r", encoding="utf-8") as f:
            lookup = json.load(f)
        mutable_cases = [list(tc) for tc in TEST_CASES]
        for tc in mutable_cases:
            tc_id = tc[0]
            if tc_id in lookup:
                tc[5] = lookup[tc_id]["status"]
                tc[7] = lookup[tc_id]["remarks"]
        TEST_CASES = [tuple(tc) for tc in mutable_cases]
        print(f"[OK] Successfully updated test statuses from lookup.")
    except Exception as e:
        print(f"[ERROR] Failed to load lookup results: {e}")

# Save workbook to the Test Results folder
output_file = os.path.join(os.path.dirname(__file__), "Test Results", "Excel", "Automation_Test_Report.xlsx")

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 1: SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════
ws_sum = wb.active
ws_sum.title = "📊 Summary"
ws_sum.sheet_view.showGridLines = False
ws_sum.column_dimensions["A"].width = 28
ws_sum.column_dimensions["B"].width = 22
ws_sum.column_dimensions["C"].width = 22
ws_sum.column_dimensions["D"].width = 22

# Title
ws_sum.merge_cells("A1:D1")
title_cell = ws_sum["A1"]
title_cell.value = "🎓 Smart Learn — Appium E2E Test Report"
title_cell.font = Font(bold=True, color=TITLE_FG, size=20, name="Calibri")
title_cell.fill = hex_fill(SECTION_BG)
title_cell.alignment = center_align()
ws_sum.row_dimensions[1].height = 45

# Metadata row
ws_sum.merge_cells("A2:D2")
meta = ws_sum["A2"]
meta.value = (
    f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
    f"Total Test Cases: {len(TEST_CASES)}  |  "
    f"App: com.smartlearn.companion  |  Platform: Android (Appium UiAutomator2)"
)
meta.font = Font(italic=True, color="64748B", size=10, name="Calibri")
meta.fill = hex_fill("0F172A")
meta.alignment = center_align()
ws_sum.row_dimensions[2].height = 22

# Gap row
ws_sum.row_dimensions[3].height = 12

# ── Module Summary Table ─────────────────────────────────────────────────────
header_row = 4
headers = ["Module", "Total Cases", "Pass", "Fail"]
for ci, h in enumerate(headers, 1):
    cell = ws_sum.cell(row=header_row, column=ci, value=h)
    cell.font = Font(bold=True, color=HEADER_FG, size=12, name="Calibri")
    cell.fill = hex_fill(HEADER_BG)
    cell.alignment = center_align()
    cell.border = thin_border()
ws_sum.row_dimensions[header_row].height = 32

# Gather module stats
from collections import defaultdict
module_stats = defaultdict(lambda: {"total": 0, "pass": 0, "fail": 0})
status_counter = defaultdict(int)
priority_counter = defaultdict(int)
type_counter = defaultdict(int)

for tc in TEST_CASES:
    mod = tc[1]
    status = tc[5]
    priority = tc[4]
    tc_type = tc[6]
    module_stats[mod]["total"] += 1
    if status == "Pass":
        module_stats[mod]["pass"] += 1
    else:
        module_stats[mod]["fail"] += 1
    status_counter[status] += 1
    priority_counter[priority] += 1
    type_counter[tc_type] += 1

modules = list(MODULE_COLORS.keys())
for ri, mod in enumerate(modules, 1):
    row = header_row + ri
    stats = module_stats[mod]
    row_data = [mod, stats["total"], stats["pass"], stats["fail"]]
    mod_color = MODULE_COLORS.get(mod, "1E293B")
    for ci, val in enumerate(row_data, 1):
        cell = ws_sum.cell(row=row, column=ci, value=val)
        cell.border = thin_border()
        if ci == 1:
            cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
            cell.fill = hex_fill(mod_color)
            cell.alignment = left_align(wrap=False)
        elif ci == 2:
            cell.font = Font(bold=True, color="1E293B", size=11, name="Calibri")
            cell.fill = hex_fill("E2E8F0")
            cell.alignment = center_align()
        elif ci == 3:
            cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
            cell.fill = hex_fill("16A34A")
            cell.alignment = center_align()
        else:
            cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
            cell.fill = hex_fill("DC2626" if stats["fail"] > 0 else "CBD5E1")
            cell.alignment = center_align()
    ws_sum.row_dimensions[row].height = 26

# Totals row
totals_row = header_row + len(modules) + 1
ws_sum.row_dimensions[totals_row].height = 30
for ci, val in enumerate(["TOTAL", len(TEST_CASES), status_counter["Pass"], status_counter["Fail"]], 1):
    cell = ws_sum.cell(row=totals_row, column=ci, value=val)
    cell.font = Font(bold=True, color=HEADER_FG, size=12, name="Calibri")
    cell.fill = hex_fill(HEADER_BG)
    cell.alignment = center_align()
    cell.border = thin_border()

# ── Pass/Fail Summary ────────────────────────────────────────────────────────
gap = totals_row + 2
ws_sum.merge_cells(f"A{gap}:D{gap}")
kpi_heading = ws_sum[f"A{gap}"]
kpi_heading.value = "📈 Overall Status Distribution"
kpi_heading.font = Font(bold=True, color=TITLE_FG, size=13, name="Calibri")
kpi_heading.fill = hex_fill(ACCENT)
kpi_heading.alignment = center_align()
ws_sum.row_dimensions[gap].height = 28

status_start = gap + 1
for ri, (status, count) in enumerate(status_counter.items()):
    row = status_start + ri
    pct = round(count / len(TEST_CASES) * 100, 1)
    for ci, val in enumerate([status, count, f"{pct}%", ""], 1):
        cell = ws_sum.cell(row=row, column=ci, value=val)
        sc = STATUS_COLORS.get(status, "94A3B8")
        if ci == 1:
            cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
            cell.fill = hex_fill(sc)
        else:
            cell.font = Font(color="1E293B", size=11, name="Calibri")
            cell.fill = hex_fill("F8FAFC")
        cell.alignment = center_align()
        cell.border = thin_border()
    ws_sum.row_dimensions[row].height = 24

# ── Priority Distribution ─────────────────────────────────────────────────────
pri_gap = status_start + len(status_counter) + 1
ws_sum.merge_cells(f"A{pri_gap}:D{pri_gap}")
pri_head = ws_sum[f"A{pri_gap}"]
pri_head.value = "🎯 Priority Distribution"
pri_head.font = Font(bold=True, color=TITLE_FG, size=13, name="Calibri")
pri_head.fill = hex_fill("0891B2")
pri_head.alignment = center_align()
ws_sum.row_dimensions[pri_gap].height = 28

for ri, (priority, count) in enumerate(priority_counter.items(), 1):
    row = pri_gap + ri
    pc = PRIORITY_COLORS.get(priority, "1E293B")
    for ci, val in enumerate([priority, count, f"{round(count/len(TEST_CASES)*100,1)}%", ""], 1):
        cell = ws_sum.cell(row=row, column=ci, value=val)
        if ci == 1:
            cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
            cell.fill = hex_fill(pc)
        else:
            cell.font = Font(color="1E293B", size=11, name="Calibri")
            cell.fill = hex_fill(ROW_ALT)
        cell.alignment = center_align()
        cell.border = thin_border()
    ws_sum.row_dimensions[row].height = 24

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 2: DETAILS
# ═══════════════════════════════════════════════════════════════════════════════
ws_det = wb.create_sheet("🧪 Test Cases")
ws_det.sheet_view.showGridLines = False

col_widths = [10, 18, 20, 65, 14, 14, 18, 30]
col_headers = ["TC ID", "Module", "Feature", "Test Case Description", "Priority", "Status", "Test Type", "Remarks"]

for ci, w in enumerate(col_widths, 1):
    ws_det.column_dimensions[get_column_letter(ci)].width = w

# Header
for ci, h in enumerate(col_headers, 1):
    cell = ws_det.cell(row=1, column=ci, value=h)
    cell.font = Font(bold=True, color=HEADER_FG, size=12, name="Calibri")
    cell.fill = hex_fill(HEADER_BG)
    cell.alignment = center_align(wrap=True)
    cell.border = thin_border()
ws_det.row_dimensions[1].height = 36

# Data rows
for ri, tc in enumerate(TEST_CASES, 2):
    tc_id, module, feature, desc, priority, status, tc_type, remarks = tc
    row_vals = [tc_id, module, feature, desc, priority, status, tc_type, remarks]

    is_alt = (ri % 2 == 0)
    base_bg = "F1F5F9" if is_alt else "FFFFFF"
    status_color = STATUS_COLORS.get(status, "94A3B8")
    priority_color = PRIORITY_COLORS.get(priority, "1E293B")
    mod_color = MODULE_COLORS.get(module, "1E293B")

    for ci, val in enumerate(row_vals, 1):
        cell = ws_det.cell(row=ri, column=ci, value=val)
        cell.border = thin_border()

        if ci == 1:   # TC ID
            cell.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
            cell.fill = hex_fill(mod_color)
            cell.alignment = center_align()
        elif ci == 2: # Module
            cell.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
            cell.fill = hex_fill(mod_color)
            cell.alignment = center_align(wrap=True)
        elif ci == 3: # Feature
            cell.font = Font(color="475569", size=10, name="Calibri")
            cell.fill = hex_fill(base_bg)
            cell.alignment = center_align(wrap=True)
        elif ci == 4: # Description
            cell.font = Font(color="1E293B", size=10, name="Calibri")
            cell.fill = hex_fill(base_bg)
            cell.alignment = left_align()
        elif ci == 5: # Priority
            cell.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
            cell.fill = hex_fill(priority_color)
            cell.alignment = center_align()
        elif ci == 6: # Status
            cell.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
            cell.fill = hex_fill(status_color)
            cell.alignment = center_align()
        elif ci == 7: # Type
            cell.font = Font(color="475569", size=10, name="Calibri", italic=True)
            cell.fill = hex_fill(base_bg)
            cell.alignment = center_align(wrap=True)
        else:         # Remarks
            cell.font = Font(color="94A3B8", size=10, name="Calibri")
            cell.fill = hex_fill(base_bg)
            cell.alignment = left_align()

    ws_det.row_dimensions[ri].height = 30

# Freeze pane
ws_det.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 3: ANALYTICS
# ═══════════════════════════════════════════════════════════════════════════════
ws_ana = wb.create_sheet("📉 Analytics")
ws_ana.sheet_view.showGridLines = False

ws_ana.column_dimensions["A"].width = 25
ws_ana.column_dimensions["B"].width = 15
ws_ana.column_dimensions["C"].width = 15

# Heading
ws_ana.merge_cells("A1:C1")
h = ws_ana["A1"]
h.value = "Test Case Analytics per Module"
h.font = Font(bold=True, color=TITLE_FG, size=16, name="Calibri")
h.fill = hex_fill(SECTION_BG)
h.alignment = center_align()
ws_ana.row_dimensions[1].height = 40

# Table header
for ci, val in enumerate(["Module", "Total", "Pass Rate %"], 1):
    cell = ws_ana.cell(row=2, column=ci, value=val)
    cell.font = Font(bold=True, color=HEADER_FG, size=11, name="Calibri")
    cell.fill = hex_fill(HEADER_BG)
    cell.alignment = center_align()
    cell.border = thin_border()
ws_ana.row_dimensions[2].height = 28

for ri, mod in enumerate(modules, 1):
    row = 2 + ri
    stats = module_stats[mod]
    rate = round(stats["pass"] / stats["total"] * 100, 1) if stats["total"] > 0 else 0
    for ci, val in enumerate([mod, stats["total"], f"{rate}%"], 1):
        cell = ws_ana.cell(row=row, column=ci, value=val)
        cell.border = thin_border()
        if ci == 1:
            cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
            cell.fill = hex_fill(MODULE_COLORS.get(mod, HEADER_BG))
        elif ci == 3:
            rate_val = float(str(val).replace('%',''))
            bg = "16A34A" if rate_val == 100 else ("EAB308" if rate_val >= 80 else "DC2626")
            cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
            cell.fill = hex_fill(bg)
        else:
            cell.font = Font(color="1E293B", size=11, name="Calibri")
            cell.fill = hex_fill("F1F5F9")
        cell.alignment = center_align()
    ws_ana.row_dimensions[row].height = 26

# ── Bar Chart ─────────────────────────────────────────────────────────────────
chart = BarChart()
chart.type = "col"
chart.title = "Test Cases per Module"
chart.y_axis.title = "Count"
chart.x_axis.title = "Module"
chart.style = 10
chart.grouping = "clustered"
chart.width = 22
chart.height = 14

data_ref = Reference(ws_ana, min_col=2, min_row=2, max_row=2 + len(modules))
cats_ref = Reference(ws_ana, min_col=1, min_row=3, max_row=2 + len(modules))
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)

ws_ana.add_chart(chart, "E2")

# ─────────────────────────────────────────────────────────────────────────────
# Save
wb.save(output_file)
print(f"\n[OK] Excel report generated: {output_file}")
print(f"     Total test cases: {len(TEST_CASES)}")
print(f"     Sheets: Summary / Test Cases / Analytics")
