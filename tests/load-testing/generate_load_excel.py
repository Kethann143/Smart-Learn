"""
generate_load_excel.py
─────────────────────────────────────────────────────────────────────────────
Smart Learn — Load Test Report Excel Generator
Reads metrics from load_test_results.json and creates a gorgeous,
professionally styled dark-themed spreadsheet with KPI blocks, detailed stats,
timeline logs, and an integrated native Excel chart for RPS vs Latency.
─────────────────────────────────────────────────────────────────────────────
"""

import os
import json
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from datetime import datetime

# ─────────────────────────────────────────────────────────────────────────────
# Styling Constants
# ─────────────────────────────────────────────────────────────────────────────

# Color Palette (Dark Theme / Cyberpunk Highlights)
COLOR_DARK_BG      = "0A0820"  # Deep dark space background
COLOR_CARD_BG      = "0D0B20"  # Card/KPI background
COLOR_ROW_ALT      = "13112E"  # Alternate row background
COLOR_PRIMARY      = "00F2FE"  # Cyan highlight
COLOR_SECONDARY    = "B152FF"  # Purple accent
COLOR_SUCCESS      = "00FF80"  # Neon green
COLOR_WARNING      = "FFAA00"  # Yellow/Orange
COLOR_DANGER       = "FF4466"  # Coral red
COLOR_TEXT_LIGHT   = "FFFFFF"  # White text
COLOR_TEXT_MUTED   = "888899"  # Gray text
COLOR_BORDER       = "2A2845"  # Gridline borders

# Fills
FILL_HEADER        = PatternFill("solid", fgColor=COLOR_DARK_BG)
FILL_ALT_ROW       = PatternFill("solid", fgColor=COLOR_ROW_ALT)
FILL_KPI           = PatternFill("solid", fgColor=COLOR_CARD_BG)
FILL_NORMAL        = PatternFill("solid", fgColor=COLOR_DARK_BG)

# Fonts
FONT_TITLE         = Font(name="Segoe UI", bold=True, size=18, color=COLOR_PRIMARY)
FONT_SECTION       = Font(name="Segoe UI", bold=True, size=12, color=COLOR_SECONDARY)
FONT_HEADER        = Font(name="Segoe UI", bold=True, size=10, color=COLOR_TEXT_LIGHT)
FONT_BODY          = Font(name="Segoe UI", size=9, color="C8C8E0")
FONT_BOLD          = Font(name="Segoe UI", bold=True, size=9, color=COLOR_TEXT_LIGHT)
FONT_VALUE_KPI     = Font(name="Segoe UI", bold=True, size=22, color=COLOR_PRIMARY)
FONT_LABEL_KPI     = Font(name="Segoe UI", bold=True, size=8, color=COLOR_TEXT_MUTED)

ALIGN_CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_RIGHT        = Alignment(horizontal="right",  vertical="center")

def thin_border(color=COLOR_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border(color=COLOR_PRIMARY):
    s = Side(style="medium", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

# ─────────────────────────────────────────────────────────────────────────────
# Report Generator
# ─────────────────────────────────────────────────────────────────────────────

def build_report():
    results_path = os.path.join(os.path.dirname(__file__), 'load_test_results.json')
    if not os.path.exists(results_path):
        print(f"❌ Error: {results_path} not found. Please run the load test first.")
        return

    with open(results_path, 'r') as f:
        data = json.load(f)

    meta = data['metadata']
    summary = data['summary']
    timeline = data['timeline']

    print("📊 Generating Excel load test report...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default active sheet

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 1: Summary Dashboard
    # ═════════════════════════════════════════════════════════════════════════
    ws_sum = wb.create_sheet("Load Test Summary")
    ws_sum.sheet_view.showGridLines = False
    ws_sum.tab_color = COLOR_PRIMARY

    # Header Row
    ws_sum.merge_cells("A1:J1")
    ws_sum["A1"] = "⚡ SMART LEARN — BACKEND BASELINE LOAD TEST REPORT"
    ws_sum["A1"].font = FONT_TITLE
    ws_sum["A1"].fill = FILL_HEADER
    ws_sum["A1"].alignment = ALIGN_CENTER
    ws_sum.row_dimensions[1].height = 40

    ws_sum.merge_cells("A2:J2")
    ws_sum["A2"] = f"Run Date: {meta['testStartTimeStr']}  |  Target Endpoint: {meta['targetUrl']}  |  Duration: {meta['plannedDurationSeconds']}s"
    ws_sum["A2"].font = Font(name="Segoe UI", italic=True, size=9, color=COLOR_TEXT_MUTED)
    ws_sum["A2"].fill = FILL_HEADER
    ws_sum["A2"].alignment = ALIGN_CENTER
    ws_sum.row_dimensions[2].height = 18

    # KPI Block spacing
    ws_sum.row_dimensions[3].height = 10

    # KPI Cards (Merged cell blocks)
    kpis = [
        ("TOTAL REQUESTS", summary['totalRequests'], "A4:B6", COLOR_PRIMARY),
        ("SUCCESS RATE", f"{summary['successRate']:.1f}%", "C4:D6", COLOR_SUCCESS),
        ("AVG LATENCY", f"{summary['avgLatency']:.1f}ms", "E4:F6", COLOR_SECONDARY),
        ("THROUGHPUT", f"{summary['averageRps']:.1f} RPS", "G4:H6", COLOR_WARNING),
        ("99th PERCENTILE", f"{summary['p99']}ms", "I4:J6", COLOR_DANGER)
    ]

    for label, val, cell_range, highlight_color in kpis:
        ws_sum.merge_cells(cell_range)
        top_left = ws_sum[cell_range.split(":")[0]]
        # Store value in it
        top_left.value = f"{label}\n{val}"
        top_left.font = Font(name="Segoe UI", bold=True, size=16, color=highlight_color)
        top_left.fill = FILL_KPI
        top_left.alignment = ALIGN_CENTER
        top_left.border = thick_border(highlight_color)

    for r in range(4, 7):
        ws_sum.row_dimensions[r].height = 25

    ws_sum.row_dimensions[7].height = 15

    # Section header
    ws_sum.merge_cells("A8:D8")
    ws_sum["A8"] = "📋 TEST RUN CONFIGURATIONS"
    ws_sum["A8"].font = FONT_SECTION
    ws_sum["A8"].alignment = ALIGN_LEFT

    ws_sum.merge_cells("F8:J8")
    ws_sum["F8"] = "📈 LATENCY PERCENTILES BREAKDOWN"
    ws_sum["F8"].font = FONT_SECTION
    ws_sum["F8"].alignment = ALIGN_LEFT
    ws_sum.row_dimensions[8].height = 22

    # Left: Configurations Table
    configs = [
        ("Target API Endpoint", meta['targetUrl']),
        ("Virtual Users (VUs)", f"{meta['concurrency']} concurrent"),
        ("Planned Duration", f"{meta['plannedDurationSeconds']} seconds"),
        ("Actual Run Duration", f"{meta['actualDurationMs'] / 1000:.2f} seconds"),
        ("Test Start Time", meta['testStartTimeStr']),
        ("Test End Time", meta['testEndTimeStr'])
    ]

    for idx, (param, value) in enumerate(configs):
        row = 9 + idx
        ws_sum.row_dimensions[row].height = 18
        alt = (idx % 2 == 1)
        bg = FILL_ALT_ROW if alt else FILL_NORMAL
        
        # Merge columns A&B, C&D
        ws_sum.merge_cells(f"A{row}:B{row}")
        ws_sum.merge_cells(f"C{row}:D{row}")
        
        c_param = ws_sum[f"A{row}"]
        c_param.value = param
        c_param.font = FONT_BOLD
        c_param.fill = bg
        c_param.alignment = ALIGN_LEFT
        c_param.border = thin_border()
        
        c_val = ws_sum[f"C{row}"]
        c_val.value = value
        c_val.font = FONT_BODY
        c_val.fill = bg
        c_val.alignment = ALIGN_LEFT
        c_val.border = thin_border()

    # Right: Latency Percentiles Table
    percentiles = [
        ("Minimum Latency", f"{summary['minLatency']} ms", COLOR_PRIMARY),
        ("50th Percentile (p50 / Median)", f"{summary['p50']} ms", COLOR_PRIMARY),
        ("90th Percentile (p90)", f"{summary['p90']} ms", COLOR_SECONDARY),
        ("95th Percentile (p95)", f"{summary['p95']} ms", COLOR_WARNING),
        ("99th Percentile (p99)", f"{summary['p99']} ms", COLOR_DANGER),
        ("Maximum Latency (Max)", f"{summary['maxLatency']} ms", COLOR_DANGER)
    ]

    for idx, (pct, value, highlight_color) in enumerate(percentiles):
        row = 9 + idx
        alt = (idx % 2 == 1)
        bg = FILL_ALT_ROW if alt else FILL_NORMAL
        
        ws_sum.merge_cells(f"F{row}:H{row}")
        ws_sum.merge_cells(f"I{row}:J{row}")
        
        c_pct = ws_sum[f"F{row}"]
        c_pct.value = pct
        c_pct.font = FONT_BOLD
        c_pct.fill = bg
        c_pct.alignment = ALIGN_LEFT
        c_pct.border = thin_border()
        
        c_val = ws_sum[f"I{row}"]
        c_val.value = value
        c_val.font = Font(name="Segoe UI", bold=True, size=9, color=highlight_color)
        c_val.fill = bg
        c_val.alignment = ALIGN_RIGHT
        c_val.border = thin_border()

    # Section spacing
    ws_sum.row_dimensions[15].height = 15

    # Description of Baseline test
    ws_sum.merge_cells("A16:J19")
    desc_text = (
        "📖 BASELINE TEST INTERPRETATION GUIDELINE:\n"
        "• Requests per Second (RPS): This tells you the capacity the server handles concurrently. "
        f"Under 100 concurrent virtual users, your API successfully handled an average of {summary['averageRps']} requests/sec.\n"
        "• Response Time: Response times below 200ms are considered excellent. "
        f"Your median response time (p50) was {summary['p50']}ms, and the overall average was {summary['avgLatency']}ms.\n"
        "• High Percentiles (p90, p99): p99 indicates the latency for the slowest 1% of users. "
        f"Here, 99% of requests completed under {summary['p99']}ms, proving excellent stability under expected concurrency."
    )
    ws_sum["A16"] = desc_text
    ws_sum["A16"].font = Font(name="Segoe UI", size=9, color="B8B8D0")
    ws_sum["A16"].fill = FILL_KPI
    ws_sum["A16"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws_sum["A16"].border = thin_border(COLOR_SECONDARY)
    ws_sum.row_dimensions[16].height = 20
    ws_sum.row_dimensions[17].height = 20
    ws_sum.row_dimensions[18].height = 20
    ws_sum.row_dimensions[19].height = 20

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 2: Timeline Progression Logs
    # ═════════════════════════════════════════════════════════════════════════
    ws_time = wb.create_sheet("Timeline Progression")
    ws_time.sheet_view.showGridLines = False
    ws_time.tab_color = COLOR_SECONDARY
    ws_time.freeze_panes = "A3"

    ws_time.merge_cells("A1:E1")
    ws_time["A1"] = "⏱️ SECOND-BY-SECOND LOAD PERFORMANCE TIMELINE LOGS"
    ws_time["A1"].font = Font(name="Segoe UI", bold=True, size=12, color=COLOR_SECONDARY)
    ws_time["A1"].fill = FILL_HEADER
    ws_time["A1"].alignment = ALIGN_CENTER
    ws_time.row_dimensions[1].height = 28

    headers = ["Timeline (Sec)", "Throughput (RPS)", "Passed Requests", "Failed Requests", "Average Latency (ms)"]
    widths = [16, 18, 16, 16, 22]
    for col, (hdr, w) in enumerate(zip(headers, widths), start=1):
        cell = ws_time.cell(row=2, column=col, value=hdr)
        cell.font = FONT_HEADER
        cell.fill = PatternFill("solid", fgColor="1A1840")
        cell.alignment = ALIGN_CENTER
        cell.border = thin_border("4444AA")
        ws_time.column_dimensions[get_column_letter(col)].width = w
    ws_time.row_dimensions[2].height = 20

    for idx, slot in enumerate(timeline):
        row = 3 + idx
        alt = (idx % 2 == 1)
        bg = FILL_ALT_ROW if alt else FILL_NORMAL
        
        vals = [
            f"Second {slot['second']}",
            slot['rps'],
            slot['passed'],
            slot['failed'],
            round(slot['avgLatency'], 1)
        ]
        
        for col, val in enumerate(vals, start=1):
            cell = ws_time.cell(row=row, column=col, value=val)
            cell.fill = bg
            cell.border = thin_border()
            cell.alignment = ALIGN_CENTER
            
            if col == 1:
                cell.font = FONT_BOLD
            elif col == 2:
                cell.font = Font(name="Segoe UI", bold=True, size=9, color=COLOR_WARNING)
            elif col == 4 and val > 0:
                cell.font = Font(name="Segoe UI", bold=True, size=9, color=COLOR_DANGER)
                cell.fill = PatternFill("solid", fgColor="330000")
            elif col == 5:
                cell.font = Font(name="Segoe UI", bold=True, size=9, color=COLOR_PRIMARY)
            else:
                cell.font = FONT_BODY
                
        ws_time.row_dimensions[row].height = 18

    # ═════════════════════════════════════════════════════════════════════════
    # NATIVE CHART GENERATION (Line Chart in Summary Dashboard)
    # ═════════════════════════════════════════════════════════════════════════
    # Create Line Chart for latency & RPS over timeline seconds
    chart = LineChart()
    chart.title = "Throughput (RPS) & Response Latency Over Time"
    chart.style = 13
    chart.y_axis.title = "Response Time / RPS"
    chart.x_axis.title = "Timeline (Seconds)"
    chart.width = 18
    chart.height = 12

    # Y-axis data: RPS (Col 2) and Latency (Col 5)
    # Reference(worksheet, min_col, min_row, max_col, max_row)
    data_ref = Reference(ws_time, min_col=2, min_row=2, max_col=2, max_row=2 + len(timeline))
    latency_ref = Reference(ws_time, min_col=5, min_row=2, max_col=5, max_row=2 + len(timeline))
    
    chart.add_data(data_ref, titles_from_data=True)
    chart.add_data(latency_ref, titles_from_data=True)

    # X-axis data: Seconds (Col 1)
    cats_ref = Reference(ws_time, min_col=1, min_row=3, max_row=2 + len(timeline))
    chart.set_categories(cats_ref)

    # Style lines
    chart.series[0].graphicalProperties.line.solidFill = COLOR_WARNING
    chart.series[1].graphicalProperties.line.solidFill = COLOR_PRIMARY

    # Add the chart to the summary sheet starting at cell A21
    ws_sum.add_chart(chart, "A21")

    # Set column widths of Summary sheet to align cleanly
    sum_widths = [14, 14, 14, 14, 14, 16, 16, 16, 14, 14]
    for idx, w in enumerate(sum_widths, start=1):
        ws_sum.column_dimensions[get_column_letter(idx)].width = w

    # Save
    output_filename = "SmartLearn_LoadTest_Report.xlsx"
    output_filepath = os.path.join(os.path.dirname(__file__), '..', output_filename)
    wb.save(output_filepath)
    print(f"✅ Excel report saved successfully to: {output_filepath}")

if __name__ == "__main__":
    build_report()
