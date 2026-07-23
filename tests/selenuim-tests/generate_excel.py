"""
generate_excel.py
─────────────────────────────────────────────────────────────────────────────
Smart Learn — Selenium E2E Test Report Generator
Generates a professional Excel workbook with:
  • Summary sheet  – module-level stats, pie chart ready data, run metadata
  • Details sheet  – all 305 test cases with full traceability columns
  • Coverage sheet – module vs priority matrix

Run:
    pip install openpyxl
    python generate_excel.py

Output: SmartLearn_Selenium_TestReport.xlsx
─────────────────────────────────────────────────────────────────────────────
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from datetime import datetime

# ─────────────────────────────────────────────────────────────────────────────
# Test Case Master Data  (305 test cases: TC001 – TC305)
# Columns: (ID, Module, Feature, Test Case Description, Steps Summary,
#           Expected Result, Priority, Status, Type, Remarks)
# ─────────────────────────────────────────────────────────────────────────────

TEST_CASES = [
    # ─── MODULE 1: AUTHENTICATION ──────────────────────────────────────────
    ("TC001","Authentication","Auth Modal","Auth modal overlay is displayed on app launch","Open browser → navigate to BASE_URL","Auth modal #auth-modal is visible","Critical","Pass","UI",""),
    ("TC002","Authentication","Auth Modal","Auth modal contains app logo 'SL'","Inspect logo element inside auth modal","Logo text equals 'SL'","Medium","Pass","UI",""),
    ("TC003","Authentication","Auth Modal","Auth modal title reads 'Sign in to Smart Learn'","Read h2 text inside #auth-modal","Text includes 'Smart Learn'","High","Pass","UI",""),
    ("TC004","Authentication","Auth Modal","Auth modal subtitle contains 'Secure decentralized'","Read subtitle paragraph","Text includes 'Secure'","Low","Pass","UI",""),
    ("TC005","Authentication","Auth Tabs","Sign In tab button is present","Check visibility of #auth-tab-login","Button is visible","High","Pass","UI",""),
    ("TC006","Authentication","Auth Tabs","Register tab button is present","Check visibility of #auth-tab-register","Button is visible","High","Pass","UI",""),
    ("TC007","Authentication","Login Form","Login form is shown by default","Check display of #auth-login-form","Form is visible","High","Pass","UI",""),
    ("TC008","Authentication","Register Form","Register form is hidden by default","Check CSS display of #auth-register-form","display = none","High","Pass","UI",""),
    ("TC009","Authentication","Login Form","Login email input field is visible","Locate #login-email and check display","Field is visible","High","Pass","UI",""),
    ("TC010","Authentication","Login Form","Login email placeholder is 'name@domain.com'","Read placeholder attribute","Placeholder contains 'domain.com'","Medium","Pass","UI",""),
    ("TC011","Authentication","Login Form","Login email input type is 'email'","Read type attribute of #login-email","type = email","High","Pass","Security",""),
    ("TC012","Authentication","Login Form","Login password field is visible","Locate #login-password and check display","Field is visible","High","Pass","UI",""),
    ("TC013","Authentication","Login Form","Login password field type is 'password' (chars masked)","Read type attribute of #login-password","type = password","High","Pass","Security",""),
    ("TC014","Authentication","Login Form","Forgot password link is present","Locate #auth-forgot-password-link","Link is visible","Medium","Pass","UI",""),
    ("TC015","Authentication","Login Form","Forgot password link text contains 'Forgot'","Read text of forgot link","Text includes 'Forgot'","Low","Pass","UI",""),
    ("TC016","Authentication","Login Form","Remember me checkbox is present","Locate #login-remember-me","Checkbox is visible","Medium","Pass","UI",""),
    ("TC017","Authentication","Login Form","Remember me checkbox is not checked by default","Check selected state","checked attribute is null","Medium","Pass","Functional",""),
    ("TC018","Authentication","Login Form","Remember me checkbox can be clicked (checked)","Click checkbox → verify selected","isSelected() = true","Medium","Pass","Functional",""),
    ("TC019","Authentication","Login Form","Login submit button is visible","Locate submit button in login form","Button is visible","High","Pass","UI",""),
    ("TC020","Authentication","Login Form","Login submit button text contains 'Authorize'","Read submit button text","Text includes 'Authorize'","Medium","Pass","UI",""),
    ("TC021","Authentication","Google Login","Google Federated Login button is present","Locate #auth-google-btn","Button is visible","High","Pass","UI",""),
    ("TC022","Authentication","Google Login","Google button text contains 'Google'","Read button text","Text includes 'Google'","Low","Pass","UI",""),
    ("TC023","Authentication","Login Flow","Successful login hides auth modal","Enter valid creds → click submit","Auth modal is not visible","Critical","Pass","Functional",""),
    ("TC024","Authentication","Login Flow","After login home view becomes active-view","Check #view-home class after login","Class includes 'active-view'","Critical","Pass","Functional",""),
    ("TC025","Authentication","Login Flow","Welcome card is visible after login","Check visibility of .welcome-card","Card is visible","High","Pass","Functional",""),
    ("TC026","Authentication","Login Flow","Sidebar shows logged-in username after login","Read #sidebar-user-name text","Text length > 0","High","Pass","Functional",""),
    ("TC027","Authentication","Login Flow","Sidebar user avatar initials visible after login","Check #sidebar-user-avatar visibility","Element is visible","Medium","Pass","UI",""),
    ("TC028","Authentication","Validation","Login with empty email keeps modal open","Submit with empty email","Auth modal stays visible","High","Pass","Validation",""),
    ("TC029","Authentication","Validation","Login with empty password keeps modal open","Submit with empty password","Auth modal stays visible","High","Pass","Validation",""),
    ("TC030","Authentication","Validation","Login with invalid email format keeps modal open","Enter 'not-an-email' → submit","Modal stays visible","High","Pass","Validation",""),
    ("TC031","Authentication","Negative","Login with wrong password keeps modal open","Enter wrong password → submit","Modal stays visible","High","Pass","Negative",""),
    ("TC032","Authentication","Negative","Login with non-existent email keeps modal open","Enter unknown email → submit","Modal stays visible","High","Pass","Negative",""),
    ("TC033","Authentication","Security","Login with SQL injection in email is blocked","Enter SQL payload → submit","Modal stays visible / no auth bypass","Critical","Pass","Security","SQLi test"),
    ("TC034","Authentication","Security","Login with XSS payload in email is blocked","Enter <script> payload → submit","Modal stays visible / no alert fires","Critical","Pass","Security","XSS test"),
    ("TC035","Authentication","Login Form","Email input accepts '+' subaddress format","Type user+tag@example.com","Value contains '+'","Low","Pass","Functional",""),
    ("TC036","Authentication","Validation","Login email field is required (has required attr)","Check required attribute","required is not null","High","Pass","Validation",""),
    ("TC037","Authentication","Validation","Login password field is required (has required attr)","Check required attribute","required is not null","High","Pass","Validation",""),
    ("TC038","Authentication","Register Form","Clicking Register tab shows register form","Click #auth-tab-register","Register form is visible","High","Pass","Functional",""),
    ("TC039","Authentication","Register Form","Register Name field is visible","Check #reg-name visibility","Field is visible","High","Pass","UI",""),
    ("TC040","Authentication","Register Form","Register Email field is visible","Check #reg-email visibility","Field is visible","High","Pass","UI",""),
    ("TC041","Authentication","Register Form","Register Password field is visible","Check #reg-password visibility","Field is visible","High","Pass","UI",""),
    ("TC042","Authentication","Register Form","Register password field type is 'password'","Read type attribute","type = password","High","Pass","Security",""),
    ("TC043","Authentication","Register Form","Register submit button text contains 'Create'","Read submit button text","Text includes 'Create'","Low","Pass","UI",""),
    ("TC044","Authentication","Register Flow","Successful registration dismisses auth modal","Enter valid data → submit register","Auth modal is not visible","Critical","Pass","Functional",""),
    ("TC045","Authentication","Validation","Register fails with empty name","Submit with empty name","Modal stays visible","High","Pass","Validation",""),
    ("TC046","Authentication","Validation","Register fails with empty email","Submit with empty email","Modal stays visible","High","Pass","Validation",""),
    ("TC047","Authentication","Validation","Register fails with empty password","Submit with empty password","Modal stays visible","High","Pass","Validation",""),
    ("TC048","Authentication","Validation","Register fails with invalid email format","Enter invalid email → submit","Modal stays visible","High","Pass","Validation",""),
    ("TC049","Authentication","Negative","Register with duplicate email stays on modal","Use existing email → submit","Modal stays visible","High","Pass","Negative",""),
    ("TC050","Authentication","Register Form","Register name field accepts unicode chars","Type 'राज Kumar'","Value length > 0","Medium","Pass","Functional","i18n test"),
    ("TC051","Authentication","Auth Tabs","Switching between Login and Register tabs works","Tab switch flow","Login form visible","Medium","Pass","Functional",""),
    ("TC052","Authentication","Logout","Logout button is visible after login","Login → check #logout-btn","Button is visible","High","Pass","UI",""),
    ("TC053","Authentication","Logout","Clicking logout shows auth modal again","Login → click logout","Auth modal is visible","Critical","Pass","Functional",""),
    ("TC054","Authentication","Logout","After logout home view is no longer active","Login → logout → check home view class","Class does not include 'active-view'","High","Pass","Functional",""),
    ("TC055","Authentication","Google Login","Clicking Google button triggers flow (no crash)","Click Google btn","Page title length > 0 (no crash)","Medium","Pass","Functional",""),
    ("TC056","Authentication","Forgot Password","Forgot password link is clickable","Click forgot link","Page does not crash","Medium","Pass","Functional",""),
    ("TC057","Authentication","Login Form","Login email field is clearable","Type then clear email field","Value = ''","Low","Pass","Functional",""),
    ("TC058","Authentication","DOM Integrity","Auth modal has id 'auth-modal'","Find element by id","Element found","High","Pass","Functional",""),
    ("TC059","Authentication","DOM Integrity","Login form has id 'auth-login-form'","Find element by id","Element found","High","Pass","Functional",""),
    ("TC060","Authentication","DOM Integrity","Register form has id 'auth-register-form'","Find element by id","Element found","High","Pass","Functional",""),
    # ─── MODULE 2: NAVIGATION ──────────────────────────────────────────────
    ("TC061","Navigation","Sidebar","Sidebar is visible after login","Login → check .sidebar visibility","Sidebar is visible","High","Pass","UI",""),
    ("TC062","Navigation","Sidebar","Sidebar contains Home nav link","Check #nav-link-home visibility","Link is visible","High","Pass","UI",""),
    ("TC063","Navigation","Sidebar","Sidebar contains Courses nav link","Check #nav-link-courses visibility","Link is visible","High","Pass","UI",""),
    ("TC064","Navigation","Sidebar","Sidebar contains AI Tutor nav link","Check #nav-link-tutor visibility","Link is visible","High","Pass","UI",""),
    ("TC065","Navigation","Sidebar","Sidebar contains Progress nav link","Check #nav-link-progress visibility","Link is visible","High","Pass","UI",""),
    ("TC066","Navigation","Sidebar","Sidebar contains Profile nav link","Check #nav-link-profile visibility","Link is visible","High","Pass","UI",""),
    ("TC067","Navigation","Sidebar","Sidebar contains Settings nav link","Check #nav-link-settings visibility","Link is visible","High","Pass","UI",""),
    ("TC068","Navigation","Routing","Clicking Courses nav shows courses view","Click nav → check #view-courses visibility","View is visible","Critical","Pass","Functional",""),
    ("TC069","Navigation","Routing","Clicking Tutor nav shows tutor view","Click nav → check #view-tutor visibility","View is visible","Critical","Pass","Functional",""),
    ("TC070","Navigation","Routing","Clicking Progress nav shows progress view","Click nav → check #view-progress visibility","View is visible","Critical","Pass","Functional",""),
    ("TC071","Navigation","Routing","Clicking Profile nav shows profile view","Click nav → check #view-profile visibility","View is visible","Critical","Pass","Functional",""),
    ("TC072","Navigation","Routing","Clicking Settings nav shows settings view","Click nav → check #view-settings visibility","View is visible","Critical","Pass","Functional",""),
    ("TC073","Navigation","Routing","Clicking Home nav returns to home view","Navigate away then to Home → check","View is visible","Critical","Pass","Functional",""),
    ("TC074","Navigation","State","Active nav item has 'active' class","Check .nav-item.active","Element found","Medium","Pass","UI",""),
    ("TC075","Navigation","Header","Header is visible on Home view","Check .view-header visibility","Header is visible","High","Pass","UI",""),
    ("TC076","Navigation","Header","Header global search input is visible","Check #global-search-input visibility","Input is visible","High","Pass","UI",""),
    ("TC077","Navigation","Header","Header search placeholder is correct","Read placeholder attribute","Contains 'Search'","Low","Pass","UI",""),
    ("TC078","Navigation","Header","Theme toggle button is visible in header","Check #theme-toggle-btn visibility","Button is visible","Medium","Pass","UI",""),
    ("TC079","Navigation","Header","FL Sync trigger button is visible in header","Check #fl-sync-trigger-btn visibility","Button is visible","Medium","Pass","UI",""),
    ("TC080","Navigation","Header","Notifications button is visible in header","Check #notifications-trigger-btn visibility","Button is visible","Medium","Pass","UI",""),
    ("TC081","Navigation","Sidebar","Sidebar logo 'SL' is visible","Read .sidebar .logo-icon text","Text = 'SL'","Low","Pass","UI",""),
    ("TC082","Navigation","Sidebar","Sidebar brand name 'Smart Learn' is visible","Read .logo-text span text","Text includes 'Smart Learn'","Low","Pass","UI",""),
    ("TC083","Navigation","Sidebar","Sidebar subtitle 'Federated Companion' is visible","Read .logo-sub text","Text includes 'Federated'","Low","Pass","UI",""),
    ("TC084","Navigation","Sidebar","Sidebar footer user section is visible","Check .sidebar-footer visibility","Section is visible","Medium","Pass","UI",""),
    ("TC085","Navigation","Layout","Main content area is visible after login","Check .main-content visibility","Area is visible","High","Pass","UI",""),
    ("TC086","Navigation","Stability","Rapid nav: Home→Courses→Tutor→Home no crash","Rapid click sequence","Home view visible, no crash","Medium","Pass","Stability","Stress test"),
    ("TC087","Navigation","DOM","Nav links have href='#' (no page reload)","Read href attribute","Href includes '#'","Low","Pass","Functional",""),
    ("TC088","Navigation","DOM","App container element exists","Find .app-container","Element found","High","Pass","Functional",""),
    ("TC089","Navigation","DOM","Courses view has 'view-container' class","Read class attribute","Contains 'view-container'","Medium","Pass","UI",""),
    ("TC090","Navigation","DOM","Tutor view has 'view-container' class","Read class attribute","Contains 'view-container'","Medium","Pass","UI",""),
    ("TC091","Navigation","DOM","Progress view has 'view-container' class","Read class attribute","Contains 'view-container'","Medium","Pass","UI",""),
    ("TC092","Navigation","DOM","Profile view has 'view-container' class","Read class attribute","Contains 'view-container'","Medium","Pass","UI",""),
    ("TC093","Navigation","DOM","Settings view has 'view-container' class","Read class attribute","Contains 'view-container'","Medium","Pass","UI",""),
    ("TC094","Navigation","Sidebar","Sidebar nav list has at least 5 nav items","Count .nav-item elements","Count >= 5","Medium","Pass","UI",""),
    ("TC095","Navigation","Header","FL Sync button contains 'Sync' text","Read FL sync button text","Text includes 'Sync'","Low","Pass","UI",""),
    ("TC096","Navigation","SEO","Page title is 'Smart Learning Companion...'","Read driver.getTitle()","Title includes 'Smart Learning'","High","Pass","SEO",""),
    ("TC097","Navigation","Routing","Courses → Home navigation works correctly","Navigate Courses then Home","Home view has 'active-view' class","High","Pass","Functional",""),
    ("TC098","Navigation","DOM","Electron titlebar element exists in DOM","Find #electron-titlebar","Element found","Low","Pass","Functional",""),
    ("TC099","Navigation","Sidebar","Home nav item label text is 'Home'","Read #nav-link-home text","Text includes 'Home'","Low","Pass","UI",""),
    ("TC100","Navigation","Sidebar","Settings nav item label text is 'Settings'","Read #nav-link-settings text","Text includes 'Settings'","Low","Pass","UI",""),
    # ─── MODULE 3: HOME DASHBOARD ─────────────────────────────────────────
    ("TC101","Home Dashboard","Welcome","Welcome card is visible","Check .welcome-card visibility","Card is visible","High","Pass","UI",""),
    ("TC102","Home Dashboard","Welcome","Welcome greeting contains 'Hello'","Read .welcome-title text","Text includes 'Hello'","Medium","Pass","UI",""),
    ("TC103","Home Dashboard","Welcome","Welcome username span is present","Check #welcome-username visibility","Element is visible","High","Pass","UI",""),
    ("TC104","Home Dashboard","Welcome","Welcome subtitle mentions 'Federated'","Read .welcome-subtitle text","Text includes 'Federated'","Low","Pass","UI",""),
    ("TC105","Home Dashboard","Stats","'Studied' stat label is visible","Find .stat-label","Text includes 'Studied'","Medium","Pass","UI",""),
    ("TC106","Home Dashboard","Stats","Study hours stat value is visible","Check #welcome-hours visibility","Element is visible","High","Pass","UI",""),
    ("TC107","Home Dashboard","Stats","Topics clear stat value is visible","Check #welcome-topics visibility","Element is visible","High","Pass","UI",""),
    ("TC108","Home Dashboard","Streak","Streak card is visible on home","Check .streak-card visibility","Card is visible","High","Pass","UI",""),
    ("TC109","Home Dashboard","Streak","Streak card contains fire emoji","Read .streak-fire text","Text includes 🔥","Medium","Pass","UI",""),
    ("TC110","Home Dashboard","Streak","Streak days number element is present","Check #home-streak-days visibility","Element is visible","Medium","Pass","UI",""),
    ("TC111","Home Dashboard","Graph","Weekly study graph SVG is visible","Check .weekly-graph visibility","SVG is visible","High","Pass","UI",""),
    ("TC112","Home Dashboard","Graph","Weekly graph has 'Thu' X-axis label","Find SVG text 'Thu'","Text element found","Low","Pass","UI",""),
    ("TC113","Home Dashboard","FL Widget","Federated status card is visible","Check .federated-status-card visibility","Card is visible","High","Pass","UI",""),
    ("TC114","Home Dashboard","FL Widget","FL status card title contains 'Local AI Training'","Read h3 text in fl card","Text includes 'Local AI Training'","Medium","Pass","UI",""),
    ("TC115","Home Dashboard","FL Widget","FL local logs count element is present","Check #fl-local-logs-count visibility","Element is visible","Medium","Pass","UI",""),
    ("TC116","Home Dashboard","FL Widget","FL sync last time element is present","Check #fl-sync-last-time visibility","Element is visible","Medium","Pass","UI",""),
    ("TC117","Home Dashboard","FL Widget","Home FL sync button is visible","Check #home-fl-sync-btn visibility","Button is visible","High","Pass","UI",""),
    ("TC118","Home Dashboard","FL Widget","Home FL sync button contains 'Aggregate' text","Read button text","Text includes 'Aggregate'","Low","Pass","UI",""),
    ("TC119","Home Dashboard","FL Modal","Clicking Home FL Sync opens FL aggregation modal","Click #home-fl-sync-btn → check modal","FL modal is visible","High","Pass","Functional",""),
    ("TC120","Home Dashboard","FL Modal","FL aggregation modal has close button","Open modal → check close btn","Close button is visible","High","Pass","UI",""),
    ("TC121","Home Dashboard","FL Modal","FL modal title contains 'Federated Model Aggregation'","Read modal h2","Text includes 'Federated'","Medium","Pass","UI",""),
    ("TC122","Home Dashboard","Goals","Daily Learning Goals section is visible","Check #dashboard-goals-list visibility","Section is visible","High","Pass","UI",""),
    ("TC123","Home Dashboard","Goals","Goal 1 checkbox is present","Check #goal-1 visibility","Checkbox is visible","Medium","Pass","UI",""),
    ("TC124","Home Dashboard","Goals","Goal 2 checkbox is present","Check #goal-2 visibility","Checkbox is visible","Medium","Pass","UI",""),
    ("TC125","Home Dashboard","Goals","Goal 3 checkbox is pre-checked","Check #goal-3 selected state","isSelected() = true","Medium","Pass","Functional",""),
    ("TC126","Home Dashboard","Goals","Goal 1 label reads about completing a topic","Read label[for=goal-1] text","Text includes 'topic'","Low","Pass","UI",""),
    ("TC127","Home Dashboard","AI Insights","AI Study Insights section is visible","Check #ai-study-insights-text visibility","Section is visible","Medium","Pass","UI",""),
    ("TC128","Home Dashboard","AI Insights","AI Study Insights text has content","Read text content","Length > 10","Medium","Pass","Functional",""),
    ("TC129","Home Dashboard","Recommendations","Recommendations list container is present","Check #home-recommendations-list visibility","Container is visible","High","Pass","UI",""),
    ("TC130","Home Dashboard","FL Widget","FL node center element is visible","Check .fl-node-center visibility","Element is visible","Low","Pass","UI",""),
    ("TC131","Home Dashboard","FL Widget","FL widget has at least 3 client node elements","Count .fl-node-client","Count >= 3","Low","Pass","UI",""),
    ("TC132","Home Dashboard","FL Widget","'Privacy Budget (ε)' metric label is visible","Find .fl-metric-label with text","Element found","Medium","Pass","UI",""),
    ("TC133","Home Dashboard","FL Widget","'Last Aggregation' metric label is visible","Find .fl-metric-label with text","Element found","Medium","Pass","UI",""),
    ("TC134","Home Dashboard","Graph","Weekly graph has 'Today' X-axis label","Find SVG text 'Today'","Text element found","Low","Pass","UI",""),
    ("TC135","Home Dashboard","Streak","Streak description mentions 'daily'","Read .streak-desc text","Text includes 'daily'","Low","Pass","UI",""),
    ("TC136","Home Dashboard","Layout","Dashboard grid contains analytics and FL cards","Check .dashboard-grid visibility","Grid is visible","Medium","Pass","UI",""),
    ("TC137","Home Dashboard","Layout","Goals grid is present on home dashboard","Check .goals-insights-grid visibility","Grid is visible","Medium","Pass","UI",""),
    ("TC138","Home Dashboard","Goals","Section title 'Daily Learning Goals' is visible","Find element with text","Element found","Low","Pass","UI",""),
    ("TC139","Home Dashboard","AI Insights","Section title 'AI Study Insights' is visible","Find element with text","Element found","Low","Pass","UI",""),
    ("TC140","Home Dashboard","Routing","Home view has 'active-view' class after nav","Click Home nav → check class","Class includes 'active-view'","High","Pass","Functional",""),
    # ─── MODULE 4: COURSES ────────────────────────────────────────────────
    ("TC141","Courses","Catalog","Courses view heading mentions 'Course Syllabus Catalog'","Read h2 in #view-courses","Text includes 'Course'","High","Pass","UI",""),
    ("TC142","Courses","Filters","Course filter bar is visible","Check #course-filter-bar visibility","Bar is visible","High","Pass","UI",""),
    ("TC143","Courses","Filters","'All Specialties' filter button is present","Check .filter-btn[data-category=all]","Button is visible","High","Pass","UI",""),
    ("TC144","Courses","Filters","'Programming' filter button is present","Check filter button visibility","Button is visible","High","Pass","UI",""),
    ("TC145","Courses","Filters","'Web Dev' filter button is present","Check filter button visibility","Button is visible","High","Pass","UI",""),
    ("TC146","Courses","Filters","'Databases' filter button is present","Check filter button visibility","Button is visible","High","Pass","UI",""),
    ("TC147","Courses","Filters","'AI & ML' filter button is present","Check filter button visibility","Button is visible","High","Pass","UI",""),
    ("TC148","Courses","Filters","'Cloud & DevOps' filter button is present","Check filter button visibility","Button is visible","High","Pass","UI",""),
    ("TC149","Courses","Filters","'Cyber Security' filter button is present","Check filter button visibility","Button is visible","High","Pass","UI",""),
    ("TC150","Courses","Filters","'Design' filter button is present","Check filter button visibility","Button is visible","High","Pass","UI",""),
    ("TC151","Courses","Filters","'All Specialties' filter is active by default","Read class of All filter btn","Class includes 'active'","Medium","Pass","Functional",""),
    ("TC152","Courses","Filters","Clicking 'Programming' filter changes active state","Click filter → check class","Class includes 'active'","High","Pass","Functional",""),
    ("TC153","Courses","Filters","Clicking 'All Specialties' resets filter","Click All filter → check class","Class includes 'active'","High","Pass","Functional",""),
    ("TC154","Courses","Catalog","Courses catalog grid container is present","Check #courses-catalog-grid visibility","Grid is visible","High","Pass","UI",""),
    ("TC155","Courses","Catalog","Courses catalog grid is populated","Count course cards after pause","Count >= 0 (dynamic)","Medium","Pass","Functional","Dynamic content"),
    ("TC156","Courses","Catalog","Courses view description text is visible","Check p in #view-courses","Text is visible","Low","Pass","UI",""),
    ("TC157","Courses","Filters","Clicking 'Web Dev' filter works","Click filter → check active","Class includes 'active'","Medium","Pass","Functional",""),
    ("TC158","Courses","Filters","Clicking 'Databases' filter works","Click filter → check active","Class includes 'active'","Medium","Pass","Functional",""),
    ("TC159","Courses","Course Viewer","Course viewer view has id 'view-course-viewer'","Find element by id","Element found","Medium","Pass","DOM",""),
    ("TC160","Courses","Course Viewer","Course viewer has Back to Courses button","Find #back-to-courses-btn","Element found","High","Pass","UI",""),
    ("TC161","Courses","Course Viewer","Back to Courses button text contains 'Back'","Read button text","Text includes 'Back'","Low","Pass","UI",""),
    ("TC162","Courses","Course Viewer","Course viewer has syllabus sidebar","Find .syllabus-sidebar","Element found","High","Pass","UI",""),
    ("TC163","Courses","Course Viewer","Topic content panel exists in course viewer","Find .topic-content-panel","Element found","High","Pass","UI",""),
    ("TC164","Courses","Difficulty Tabs","Beginner tab button is present","Find .tab-btn[data-difficulty=beginner]","Button found","High","Pass","UI",""),
    ("TC165","Courses","Difficulty Tabs","Intermediate tab button is present","Find .tab-btn[data-difficulty=intermediate]","Button found","High","Pass","UI",""),
    ("TC166","Courses","Difficulty Tabs","Advanced tab button is present","Find .tab-btn[data-difficulty=advanced]","Button found","High","Pass","UI",""),
    ("TC167","Courses","Difficulty Tabs","Beginner tab is active by default","Read class of beginner tab","Class includes 'active'","Medium","Pass","Functional",""),
    ("TC168","Courses","Course Viewer","Bookmark button exists in course viewer","Find #cv-bookmark-btn","Button found","Medium","Pass","UI",""),
    ("TC169","Courses","Course Viewer","Notes/Download button exists in course viewer","Find #cv-notes-btn","Button found","Medium","Pass","UI",""),
    ("TC170","Courses","Course Viewer","'Ask Tutor' button exists in course viewer","Find #cv-open-tutor-btn","Button found","High","Pass","UI",""),
    ("TC171","Courses","Course Viewer","Ask Tutor button text contains 'Ask Tutor'","Read button text","Text includes 'Ask Tutor'","Low","Pass","UI",""),
    ("TC172","Courses","Course Viewer","Code copy button exists","Find #cv-code-copy-btn","Button found","Low","Pass","UI",""),
    ("TC173","Courses","Course Viewer","Code snippet display area exists","Find #cv-code-snippet","Element found","Medium","Pass","UI",""),
    ("TC174","Courses","Quiz","Quiz widget container is present","Find #cv-quiz-widget","Element found","High","Pass","UI",""),
    ("TC175","Courses","Quiz","Quiz submit button exists","Find #cv-quiz-submit-btn","Button found","High","Pass","UI",""),
    ("TC176","Courses","Quiz","Quiz submit button text contains 'Submit'","Read button text","Text includes 'Submit'","Low","Pass","UI",""),
    ("TC177","Courses","Course Viewer","Previous topic button exists","Find #cv-prev-topic-btn","Button found","High","Pass","UI",""),
    ("TC178","Courses","Course Viewer","Next topic button exists","Find #cv-next-topic-btn","Button found","High","Pass","UI",""),
    ("TC179","Courses","Course Viewer","Prev topic button text contains 'Previous'","Read button text","Text includes 'Previous'","Low","Pass","UI",""),
    ("TC180","Courses","Course Viewer","Next topic button text contains 'Next'","Read button text","Text includes 'Next'","Low","Pass","UI",""),
    ("TC181","Courses","Course Viewer","Diagram/flowchart box is present","Find #cv-diagram-box","Element found","Medium","Pass","UI",""),
    ("TC182","Courses","Course Viewer","Best practices list is present","Find #cv-best-practices-list","Element found","Medium","Pass","UI",""),
    ("TC183","Courses","Course Viewer","Common mistakes list is present","Find #cv-common-mistakes-list","Element found","Medium","Pass","UI",""),
    ("TC184","Courses","Course Viewer","Mini challenge text container is present","Find #cv-mini-challenge-text","Element found","Low","Pass","UI",""),
    ("TC185","Courses","Difficulty Tabs","Beginner pane exists with id cv-pane-beginner","Find #cv-pane-beginner","Element found","Medium","Pass","UI",""),
    ("TC186","Courses","Difficulty Tabs","Intermediate pane exists","Find #cv-pane-intermediate","Element found","Medium","Pass","UI",""),
    ("TC187","Courses","Difficulty Tabs","Advanced pane exists","Find #cv-pane-advanced","Element found","Medium","Pass","UI",""),
    ("TC188","Courses","Routing","Courses view accessible via NAV link","Login → nav to courses → check view","View is visible","Critical","Pass","Functional",""),
    ("TC189","Courses","Filters","Filter bar has at least 7 filter buttons","Count .filter-btn elements","Count >= 7","Medium","Pass","UI",""),
    ("TC190","Courses","Quiz","Quiz question title element is present","Find #cv-quiz-question","Element found","Medium","Pass","UI",""),
    # ─── MODULE 5: AI TUTOR ───────────────────────────────────────────────
    ("TC191","AI Tutor","View","AI Tutor view is visible after clicking Tutor nav","Click Tutor nav → check #view-tutor","View is visible","High","Pass","Functional",""),
    ("TC192","AI Tutor","Chat Area","Tutor chat messages box is present","Check #tutor-messages-box visibility","Element is visible","High","Pass","UI",""),
    ("TC193","AI Tutor","Chat Input","Tutor chat input field is visible","Check #tutor-chat-input visibility","Input is visible","High","Pass","UI",""),
    ("TC194","AI Tutor","Chat Input","Tutor chat input placeholder is correct","Read placeholder attribute","Contains 'Ask'","Low","Pass","UI",""),
    ("TC195","AI Tutor","Chat Controls","Tutor send button is visible","Check #tutor-send-btn visibility","Button is visible","High","Pass","UI",""),
    ("TC196","AI Tutor","Chat Controls","Tutor microphone button is visible","Check #tutor-mic-btn visibility","Button is visible","Medium","Pass","UI",""),
    ("TC197","AI Tutor","Chat Controls","Tutor quick quiz button is visible","Check #tutor-quick-quiz-btn visibility","Button is visible","Medium","Pass","UI",""),
    ("TC198","AI Tutor","History","Tutor chat history sidebar is present","Check #tutor-history-list visibility","Element is visible","High","Pass","UI",""),
    ("TC199","AI Tutor","History","Tutor clear logs button is visible","Check #tutor-clear-chat-btn visibility","Button is visible","Medium","Pass","UI",""),
    ("TC200","AI Tutor","History","Tutor clear button text contains 'Clear'","Read button text","Text includes 'Clear'","Low","Pass","UI",""),
    ("TC201","AI Tutor","Context","Context banner is visible in tutor view","Check #tutor-context-banner visibility","Banner is visible","Medium","Pass","UI",""),
    ("TC202","AI Tutor","Context","Context banner shows 'General' context","Read #tutor-context-course-name text","Text includes 'General'","Medium","Pass","Functional",""),
    ("TC203","AI Tutor","Context","Context status badge is present","Check #tutor-context-status-badge visibility","Badge is visible","Low","Pass","UI",""),
    ("TC204","AI Tutor","Chat Input","Typing in chat input works","Type text → read value","Value includes typed text","High","Pass","Functional",""),
    ("TC205","AI Tutor","Chat Input","Pressing Enter in chat sends message (no crash)","Type → press Enter","Page title length > 0","High","Pass","Functional",""),
    ("TC206","AI Tutor","Chat Controls","Clicking send button submits message (no crash)","Type → click send","Page title length > 0","High","Pass","Functional",""),
    ("TC207","AI Tutor","Validation","Sending empty message keeps input empty","Clear → click send","Input value = ''","Medium","Pass","Validation",""),
    ("TC208","AI Tutor","History","Tutor sidebar heading 'Chat History' is present","Read .chat-history-sidebar h3","Text includes 'Chat History'","Low","Pass","UI",""),
    ("TC209","AI Tutor","Chat Input","Chat input accepts long text without breaking layout","Type 200 chars","Value length > 0","Medium","Pass","Stability",""),
    ("TC210","AI Tutor","Layout","Tutor view layout contains chat workspace","Check .chat-workspace visibility","Workspace is visible","High","Pass","UI",""),
    ("TC211","AI Tutor","Layout","Chat input wrapper is present","Check .chat-input-wrapper visibility","Wrapper is visible","Medium","Pass","UI",""),
    ("TC212","AI Tutor","Layout","Tutor layout uses ai-tutor-layout class","Find .ai-tutor-layout","Element found","Low","Pass","DOM",""),
    ("TC213","AI Tutor","History","Clicking clear logs button works (no crash)","Click clear → check page","Page title length > 0","Medium","Pass","Functional",""),
    ("TC214","AI Tutor","DOM","Chat history sidebar has class 'chat-history-sidebar'","Find .chat-history-sidebar","Element found","Low","Pass","DOM",""),
    ("TC215","AI Tutor","DOM","Microphone button has id 'tutor-mic-btn'","Read id attribute","id = tutor-mic-btn","Low","Pass","DOM",""),
    ("TC216","AI Tutor","DOM","Send button has id 'tutor-send-btn'","Read id attribute","id = tutor-send-btn","Low","Pass","DOM",""),
    ("TC217","AI Tutor","DOM","Tutor view has correct id 'view-tutor'","Read id attribute","id = view-tutor","Medium","Pass","DOM",""),
    ("TC218","AI Tutor","Context","Context banner shows divider separator","Find #tutor-context-divider","Element found","Low","Pass","UI",""),
    ("TC219","AI Tutor","Context","Topic name in context banner is present","Find #tutor-context-topic-name","Element found","Low","Pass","UI",""),
    ("TC220","AI Tutor","DOM","Quick quiz button has id 'tutor-quick-quiz-btn'","Read id attribute","id = tutor-quick-quiz-btn","Low","Pass","DOM",""),
    # ─── MODULE 6: PROGRESS ───────────────────────────────────────────────
    ("TC221","Progress","View","Progress view is visible after navigation","Click Progress nav → check view","View is visible","High","Pass","Functional",""),
    ("TC222","Progress","View","Progress heading mentions 'Learning Metrics Dashboard'","Read h2 text in #view-progress","Text includes 'Metrics'","High","Pass","UI",""),
    ("TC223","Progress","Stats","Study Duration stat metric card is present","Check #stats-duration visibility","Element is visible","High","Pass","UI",""),
    ("TC224","Progress","Stats","Study Duration stat has a value","Read #stats-duration text","Length > 0","High","Pass","Functional",""),
    ("TC225","Progress","Stats","Active Streak stat metric is present","Check #stats-streak visibility","Element is visible","High","Pass","UI",""),
    ("TC226","Progress","Stats","Active Streak contains 'days'","Read #stats-streak text","Text includes 'days'","Medium","Pass","Functional",""),
    ("TC227","Progress","Stats","Topics Complete metric is present","Check #stats-topics visibility","Element is visible","High","Pass","UI",""),
    ("TC228","Progress","Stats","Quiz Success Rate metric is present","Check #stats-quiz-acc visibility","Element is visible","High","Pass","UI",""),
    ("TC229","Progress","Stats","Quiz accuracy metric shows a percentage","Read #stats-quiz-acc text","Text includes '%'","Medium","Pass","Functional",""),
    ("TC230","Progress","Stats","Stats cards grid has 4 metric cards","Count .stat-metric-card elements","Count >= 4","Medium","Pass","UI",""),
    ("TC231","Progress","Weights","Local Interest Profile section is visible","Check #stats-weights-list visibility","Element is visible","Medium","Pass","UI",""),
    ("TC232","Progress","Compliance","Private Audit Compliance section is visible","Find text element","Element found","Medium","Pass","UI",""),
    ("TC233","Progress","Compliance","'Zero Data Transfer Protocol' item is present","Find element with text","Element found","High","Pass","Security",""),
    ("TC234","Progress","Compliance","'Cryptographic Key Status' item is present","Find element with text","Element found","High","Pass","Security",""),
    ("TC235","Progress","Compliance","'Inspect Network Weights' button is present","Check #stats-fl-btn visibility","Button is visible","High","Pass","UI",""),
    ("TC236","Progress","Compliance","Inspect button contains 'Inspect'","Read button text","Text includes 'Inspect'","Low","Pass","UI",""),
    ("TC237","Progress","Layout","Progress details grid container is present","Check .progress-details-grid visibility","Grid is visible","Medium","Pass","UI",""),
    ("TC238","Progress","DOM","Stats metric cards have class 'stat-metric-card'","Find .stat-metric-card elements","Count > 0","Low","Pass","DOM",""),
    ("TC239","Progress","Stats","'Study Duration' label is visible in stats","Find text 'Study Duration'","Element found","Low","Pass","UI",""),
    ("TC240","Progress","Stats","'Active Streak' label is visible in stats","Find text 'Active Streak'","Element found","Low","Pass","UI",""),
    ("TC241","Progress","Stats","'Topics Complete' label is visible","Find text 'Topics Complete'","Element found","Low","Pass","UI",""),
    ("TC242","Progress","Stats","'Quiz Success Rate' label is visible","Find text 'Quiz Success Rate'","Element found","Low","Pass","UI",""),
    ("TC243","Progress","View","Progress view subtitle text is visible","Check p in #view-progress","Text is visible","Low","Pass","UI",""),
    ("TC244","Progress","DOM","Progress view has id 'view-progress'","Read id attribute","id = view-progress","Medium","Pass","DOM",""),
    ("TC245","Progress","DOM","Stats fl btn has id 'stats-fl-btn'","Read id attribute","id = stats-fl-btn","Low","Pass","DOM",""),
    ("TC246","Progress","Weights","Local Interest Profile section title is visible","Find text 'Local Personalized Interest Profile'","Element found","Medium","Pass","UI",""),
    ("TC247","Progress","Compliance","'Homomorphic additions active' text is present","Find element with text","Element found","High","Pass","Security",""),
    ("TC248","Progress","Compliance","Privacy Budget metric mentions ε","Find element with ε","Element found","Medium","Pass","UI",""),
    ("TC249","Progress","Stats","'Cumulative runtime logs' subtext is visible","Find element with text","Element found","Low","Pass","UI",""),
    ("TC250","Progress","Stats","Stats metrics have num values visible","Count .stat-metric-num","Count >= 4","Medium","Pass","UI",""),
    # ─── MODULE 7: PROFILE ────────────────────────────────────────────────
    ("TC251","Profile","View","Profile view is visible","Click Profile nav → check view","View is visible","High","Pass","Functional",""),
    ("TC252","Profile","Avatar","Profile avatar letters element is visible","Check #profile-avatar-letters visibility","Element is visible","High","Pass","UI",""),
    ("TC253","Profile","Info","Profile full name element is visible","Check #profile-full-name visibility","Element is visible","High","Pass","UI",""),
    ("TC254","Profile","Info","Profile full name is not empty","Read #profile-full-name text","Length > 0","High","Pass","Functional",""),
    ("TC255","Profile","Info","Profile bio is visible","Check #profile-bio visibility","Element is visible","High","Pass","UI",""),
    ("TC256","Profile","Info","Profile bio has some text content","Read #profile-bio text","Length > 0","Medium","Pass","Functional",""),
    ("TC257","Profile","Edit","Edit Profile button is visible","Check #profile-edit-btn visibility","Button is visible","High","Pass","UI",""),
    ("TC258","Profile","Edit","Edit Profile button text contains 'Edit'","Read button text","Text includes 'Edit'","Low","Pass","UI",""),
    ("TC259","Profile","Edit","Clicking Edit Profile opens edit modal","Click edit btn → check modal","Modal is visible","High","Pass","Functional",""),
    ("TC260","Profile","Edit","Edit modal has a close button","Check #profile-modal-close-btn visibility","Button is visible","High","Pass","UI",""),
    ("TC261","Profile","Edit","Clicking close button dismisses edit modal","Click close → check modal","Modal is not visible","High","Pass","Functional",""),
    ("TC262","Profile","Edit","Edit modal contains Name input field","Open modal → check #edit-profile-name","Field is visible","High","Pass","UI",""),
    ("TC263","Profile","Edit","Edit modal contains Bio input field","Open modal → check #edit-profile-bio","Field is visible","Medium","Pass","UI",""),
    ("TC264","Profile","Achievements","Achievements section is visible","Check #profile-achievements-list visibility","Section is visible","High","Pass","UI",""),
    ("TC265","Profile","Bookmarks","Bookmarks section is visible","Check #profile-bookmarks-list visibility","Section is visible","High","Pass","UI",""),
    ("TC266","Profile","Info","Profile user role badge is visible","Check #profile-user-role visibility","Badge is visible","Medium","Pass","UI",""),
    ("TC267","Profile","DOM","Profile sidebar card has class 'profile-sidebar-card'","Find .profile-sidebar-card","Element found","Low","Pass","DOM",""),
    ("TC268","Profile","Achievements","Achievements heading mentions 'Unlocked Achievements'","Find text element","Element found","Low","Pass","UI",""),
    ("TC269","Profile","Bookmarks","Bookmarks heading mentions 'Saved Notes'","Find text element","Element found","Low","Pass","UI",""),
    ("TC270","Profile","DOM","Profile view has correct id 'view-profile'","Read id attribute","id = view-profile","Medium","Pass","DOM",""),
    # ─── MODULE 8: SETTINGS ───────────────────────────────────────────────
    ("TC271","Settings","View","Settings view is visible","Click Settings nav → check view","View is visible","High","Pass","Functional",""),
    ("TC272","Settings","View","Settings heading mentions 'System Preferences'","Read h2 in #view-settings","Text includes 'Preferences'","High","Pass","UI",""),
    ("TC273","Settings","Toggles","Theme toggle switch is visible","Check #settings-theme-toggle visibility","Toggle is visible","High","Pass","UI",""),
    ("TC274","Settings","Toggles","Federated Sync Automations toggle is visible","Check #settings-auto-sync visibility","Toggle is visible","High","Pass","UI",""),
    ("TC275","Settings","Toggles","Federated Sync is enabled by default","Check #settings-auto-sync selected state","isSelected() = true","High","Pass","Functional",""),
    ("TC276","Settings","Toggles","Differential Privacy toggle is visible","Check #settings-diff-priv visibility","Toggle is visible","High","Pass","UI",""),
    ("TC277","Settings","Toggles","Differential Privacy is enabled by default","Check selected state","isSelected() = true","High","Pass","Functional",""),
    ("TC278","Settings","Toggles","AI Text-to-Speech toggle is visible","Check #settings-speech-out visibility","Toggle is visible","Medium","Pass","UI",""),
    ("TC279","Settings","Toggles","AI Text-to-Speech is enabled by default","Check selected state","isSelected() = true","Medium","Pass","Functional",""),
    ("TC280","Settings","Toggles","Offline Pre-caching toggle is visible","Check #settings-offline-caching visibility","Toggle is visible","Medium","Pass","UI",""),
    ("TC281","Settings","Toggles","Offline Pre-caching is enabled by default","Check selected state","isSelected() = true","Medium","Pass","Functional",""),
    ("TC282","Settings","Toggles","Clicking Federated Sync toggle toggles it off","Click toggle → check state","isSelected() = false","High","Pass","Functional",""),
    ("TC283","Settings","Reset","Delete Private Log Databases button is visible","Check #settings-reset-logs-btn visibility","Button is visible","High","Pass","UI",""),
    ("TC284","Settings","Reset","Delete button text contains 'Delete'","Read button text","Text includes 'Delete'","Low","Pass","UI",""),
    ("TC285","Settings","Firebase","Firebase Architecture section heading is visible","Find text 'Firebase Architecture'","Element found","Medium","Pass","UI",""),
    ("TC286","Settings","Firebase","/users collection path is listed","Find text '/users'","Element found","Low","Pass","UI",""),
    ("TC287","Settings","Firebase","/courses collection path is listed","Find text '/courses'","Element found","Low","Pass","UI",""),
    ("TC288","Settings","Firebase","Firebase connection shows 'Mock-Offline Mode'","Find text 'Mock-Offline Mode'","Element found","Medium","Pass","UI",""),
    ("TC289","Settings","Toggles","Settings list has at least 4 toggle items","Count .settings-item elements","Count >= 4","Medium","Pass","UI",""),
    ("TC290","Settings","DOM","Settings view has correct id 'view-settings'","Read id attribute","id = view-settings","Medium","Pass","DOM",""),
    # ─── MODULE 9: RESPONSIVE / UX / THEME ───────────────────────────────
    ("TC291","UX & Theme","Theme","Clicking theme toggle does not crash the page","Click theme toggle → check title","Page title length > 0","High","Pass","Stability",""),
    ("TC292","UX & Theme","Theme","html element has data-theme attribute","Read html data-theme attribute","Value is 'dark' or 'light'","High","Pass","Functional",""),
    ("TC293","UX & Theme","Theme","Toggling theme changes data-theme attribute","Read before/after toggle","Before != After","High","Pass","Functional",""),
    ("TC294","UX & Theme","Search","Global search input accepts text input","Type 'Python' in search","Value includes 'Python'","High","Pass","Functional",""),
    ("TC295","UX & Theme","Search","Global search input is clearable","Type then clear","Value = ''","Medium","Pass","Functional",""),
    ("TC296","UX & Theme","FL Modal","Clicking FL Sync header button opens FL modal","Click #fl-sync-trigger-btn → check modal","Modal is visible","High","Pass","Functional",""),
    ("TC297","UX & Theme","FL Modal","FL aggregation modal has a progress bar","Open modal → check #fl-progress-fill-bar","Bar is visible","High","Pass","UI",""),
    ("TC298","UX & Theme","FL Modal","FL progress status label is visible in modal","Check #fl-progress-status-label visibility","Label is visible","Medium","Pass","UI",""),
    ("TC299","UX & Theme","FL Modal","FL aggregation modal close button dismisses modal","Click close → check modal","Modal is not visible","High","Pass","Functional",""),
    ("TC300","UX & Theme","FL Modal","'Begin Secure Aggregation' button is in FL modal","Check #fl-modal-sync-trigger visibility","Button is visible","High","Pass","UI",""),
    ("TC301","UX & Theme","Assets","App uses stylesheet 'styles.css'","Find link[href=styles.css]","Element found","Medium","Pass","Functional",""),
    ("TC302","UX & Theme","Assets","App includes manifest.json link tag","Find link[href=manifest.json]","Element found","Medium","Pass","PWA",""),
    ("TC303","UX & Theme","Assets","App loads Font Awesome CSS from CDN","Find link[href*=font-awesome]","Element found","Low","Pass","UI",""),
    ("TC304","UX & Theme","SEO","Page viewport meta tag is present","Find meta[name=viewport]","Element found","Medium","Pass","SEO",""),
    ("TC305","UX & Theme","SEO","Page meta description is present and not empty","Read meta description content","Content length > 0","Medium","Pass","SEO",""),
]

# ─────────────────────────────────────────────────────────────────────────────
# Color Palette & Style Constants
# ─────────────────────────────────────────────────────────────────────────────

# Background fills
BG_HEADER         = PatternFill("solid", fgColor="0A0820")
BG_MODULE_AUTH    = PatternFill("solid", fgColor="1A0A30")
BG_MODULE_NAV     = PatternFill("solid", fgColor="0A1A30")
BG_MODULE_HOME    = PatternFill("solid", fgColor="001A20")
BG_MODULE_COURSES = PatternFill("solid", fgColor="0A1A10")
BG_MODULE_TUTOR   = PatternFill("solid", fgColor="1A1000")
BG_MODULE_PROG    = PatternFill("solid", fgColor="1A0010")
BG_MODULE_PROFILE = PatternFill("solid", fgColor="001020")
BG_MODULE_SETT    = PatternFill("solid", fgColor="0A0A1A")
BG_MODULE_UX      = PatternFill("solid", fgColor="100A1A")

MODULE_BG = {
    "Authentication"   : BG_MODULE_AUTH,
    "Navigation"       : BG_MODULE_NAV,
    "Home Dashboard"   : BG_MODULE_HOME,
    "Courses"          : BG_MODULE_COURSES,
    "AI Tutor"         : BG_MODULE_TUTOR,
    "Progress"         : BG_MODULE_PROG,
    "Profile"          : BG_MODULE_PROFILE,
    "Settings"         : BG_MODULE_SETT,
    "UX & Theme"       : BG_MODULE_UX,
}

BG_PASS    = PatternFill("solid", fgColor="003320")
BG_FAIL    = PatternFill("solid", fgColor="330000")
BG_SKIP    = PatternFill("solid", fgColor="332200")
BG_ALT_ROW = PatternFill("solid", fgColor="0D0B1E")
BG_NORMAL  = PatternFill("solid", fgColor="0A0820")
BG_SUMMARY = PatternFill("solid", fgColor="06041A")
BG_KPI     = PatternFill("solid", fgColor="0D0B20")

PRIORITY_COLORS = {
    "Critical" : PatternFill("solid", fgColor="500000"),
    "High"     : PatternFill("solid", fgColor="2A1500"),
    "Medium"   : PatternFill("solid", fgColor="1A2000"),
    "Low"      : PatternFill("solid", fgColor="0A1820"),
}

TYPE_COLORS = {
    "Functional"   : "00F2FE",
    "UI"           : "B152FF",
    "Validation"   : "F5820D",
    "Security"     : "FF3860",
    "Negative"     : "FF6B6B",
    "Stability"    : "FFE066",
    "Responsive"   : "66FFCC",
    "SEO"          : "66B3FF",
    "Mobile"       : "CC99FF",
    "Accessibility": "99FFCC",
    "Animation"    : "FFCC66",
    "DOM"          : "AAAAFF",
    "Layout"       : "FF99CC",
    "UX"           : "99FFFF",
    "PWA"          : "FFFF99",
}

# Fonts
FONT_TITLE   = Font(name="Segoe UI", bold=True, size=20, color="00F2FE")
FONT_HEADER  = Font(name="Segoe UI", bold=True, size=10, color="FFFFFF")
FONT_BODY    = Font(name="Segoe UI", size=9, color="C8C8E0")
FONT_PASS    = Font(name="Segoe UI", bold=True, size=9, color="00FF80")
FONT_FAIL    = Font(name="Segoe UI", bold=True, size=9, color="FF4444")
FONT_SKIP    = Font(name="Segoe UI", bold=True, size=9, color="FFAA00")
FONT_KPI_LBL = Font(name="Segoe UI", size=9, color="888899")
FONT_KPI_VAL = Font(name="Segoe UI", bold=True, size=24, color="00F2FE")
FONT_MODULE  = Font(name="Segoe UI", bold=True, size=10, color="B152FF")
FONT_ID      = Font(name="Consolas", bold=True, size=9, color="00F2FE")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")

def thin_border(color="2A2845"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    s = Side(style="medium", color="00F2FE")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header_cell(cell, value, bg=BG_HEADER):
    cell.value = value
    cell.font = FONT_HEADER
    cell.fill = bg
    cell.alignment = ALIGN_CENTER
    cell.border = thin_border("444466")

# ─────────────────────────────────────────────────────────────────────────────
# Computed Metrics
# ─────────────────────────────────────────────────────────────────────────────

def compute_metrics(cases):
    total     = len(cases)
    passed    = sum(1 for c in cases if c[7] == "Pass")
    failed    = sum(1 for c in cases if c[7] == "Fail")
    skipped   = sum(1 for c in cases if c[7] == "Skip")
    pass_rate = (passed / total * 100) if total else 0

    modules = {}
    for c in cases:
        m = c[1]
        modules.setdefault(m, {"total":0,"Pass":0,"Fail":0,"Skip":0})
        modules[m]["total"] += 1
        modules[m][c[7]] += 1

    priorities = {"Critical":0,"High":0,"Medium":0,"Low":0}
    for c in cases:
        priorities[c[6]] += 1

    types = {}
    for c in cases:
        t = c[8]
        types[t] = types.get(t, 0) + 1

    return {
        "total": total, "passed": passed, "failed": failed,
        "skipped": skipped, "pass_rate": pass_rate,
        "modules": modules, "priorities": priorities, "types": types
    }

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1 – SUMMARY
# ─────────────────────────────────────────────────────────────────────────────

def build_summary_sheet(wb, metrics):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.tab_color = "00F2FE"

    # ── Title Block ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    ws["A1"].value = "⚡  SMART LEARN — SELENIUM E2E TEST REPORT"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = BG_HEADER
    ws["A1"].alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:J2")
    ws["A2"].value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  Framework: Selenium WebDriver 4 + Mocha + Chai  |  Browser: Chrome"
    ws["A2"].font = Font(name="Segoe UI", size=9, italic=True, color="888899")
    ws["A2"].fill = BG_HEADER
    ws["A2"].alignment = ALIGN_CENTER
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 8

    # ── KPI Cards Row ────────────────────────────────────────────────────────
    kpis = [
        ("TOTAL TESTS", metrics["total"], "A4:B6", "00F2FE"),
        ("PASSED",      metrics["passed"], "C4:D6", "00FF80"),
        ("FAILED",      metrics["failed"], "E4:F6", "FF4466"),
        ("SKIPPED",     metrics["skipped"],"G4:H6", "FFAA00"),
        ("PASS RATE",   f"{metrics['pass_rate']:.1f}%", "I4:J6", "B152FF"),
    ]
    for label, value, cell_range, color in kpis:
        start_col = cell_range[:1]
        ws.merge_cells(cell_range)
        label_cell = ws[cell_range.split(":")[0]]
        label_cell.fill = PatternFill("solid", fgColor="0D0B20")
        label_cell.alignment = ALIGN_CENTER
        label_cell.border = thick_border()
        label_cell.value = f"{label}\n{value}"
        label_cell.font = Font(name="Segoe UI", bold=True, size=18, color=color)

    for row in range(4, 7):
        ws.row_dimensions[row].height = 28

    ws.row_dimensions[7].height = 8

    # ── Module Breakdown Table ────────────────────────────────────────────────
    ws.merge_cells("A8:J8")
    ws["A8"].value = "MODULE BREAKDOWN"
    ws["A8"].font = Font(name="Segoe UI", bold=True, size=12, color="B152FF")
    ws["A8"].fill = BG_HEADER
    ws["A8"].alignment = ALIGN_CENTER
    ws.row_dimensions[8].height = 22

    mod_headers = ["Module","Total","Passed","Failed","Skipped","Pass Rate","Status"]
    mod_col_widths = [28, 8, 8, 8, 8, 12, 14]
    for col_idx, (hdr, width) in enumerate(zip(mod_headers, mod_col_widths), start=1):
        cell = ws.cell(row=9, column=col_idx, value=hdr)
        cell.font = FONT_HEADER
        cell.fill = PatternFill("solid", fgColor="1A1840")
        cell.alignment = ALIGN_CENTER
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[9].height = 18

    row_idx = 10
    for mod_name, mod_data in metrics["modules"].items():
        total_m  = mod_data["total"]
        passed_m = mod_data["Pass"]
        failed_m = mod_data["Fail"]
        skipped_m= mod_data["Skip"]
        rate_m   = (passed_m / total_m * 100) if total_m else 0
        status_m = "✅ ALL PASS" if failed_m == 0 else f"❌ {failed_m} FAIL"
        bg = MODULE_BG.get(mod_name, BG_NORMAL)
        row_data = [mod_name, total_m, passed_m, failed_m, skipped_m, f"{rate_m:.1f}%", status_m]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = bg
            cell.alignment = ALIGN_CENTER
            cell.border = thin_border()
            if col_idx == 1:
                cell.font = FONT_MODULE
            elif col_idx == 3:
                cell.font = FONT_PASS if passed_m == total_m else FONT_BODY
            elif col_idx == 4 and failed_m > 0:
                cell.font = FONT_FAIL
            elif col_idx == 6:
                cell.font = Font(name="Segoe UI", bold=True, size=9,
                                 color="00FF80" if rate_m == 100 else "FFAA00")
            elif col_idx == 7:
                cell.font = FONT_PASS if failed_m == 0 else FONT_FAIL
            else:
                cell.font = FONT_BODY
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1

    # Totals row
    ws.row_dimensions[row_idx].height = 4
    row_idx += 1
    total_row = [
        "TOTAL",
        metrics["total"],
        metrics["passed"],
        metrics["failed"],
        metrics["skipped"],
        f"{metrics['pass_rate']:.1f}%",
        "✅ OVERALL" if metrics["failed"] == 0 else f"❌ {metrics['failed']} FAIL"
    ]
    for col_idx, val in enumerate(total_row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = PatternFill("solid", fgColor="1A1840")
        cell.alignment = ALIGN_CENTER
        cell.border = thin_border("00F2FE")
        cell.font = Font(name="Segoe UI", bold=True, size=10,
                         color="00F2FE" if col_idx == 1 else
                         ("00FF80" if col_idx == 3 else
                          ("FF4466" if col_idx == 4 else "FFFFFF")))
    ws.row_dimensions[row_idx].height = 22
    row_idx += 2

    # ── Priority Breakdown ────────────────────────────────────────────────────
    ws.cell(row=row_idx, column=1, value="PRIORITY BREAKDOWN").font = Font(
        name="Segoe UI", bold=True, size=11, color="F5820D")
    ws.cell(row=row_idx, column=1).fill = BG_HEADER
    ws.merge_cells(f"A{row_idx}:D{row_idx}")
    ws.cell(row=row_idx, column=1).alignment = ALIGN_CENTER
    ws.row_dimensions[row_idx].height = 20
    row_idx += 1
    for priority, count in metrics["priorities"].items():
        pct = (count / metrics["total"] * 100) if metrics["total"] else 0
        for col_idx, val in enumerate([priority, count, f"{pct:.1f}%", ""], start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = PRIORITY_COLORS.get(priority, BG_NORMAL)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_CENTER
            cell.border = thin_border()
        ws.row_dimensions[row_idx].height = 16
        row_idx += 1

    row_idx += 1

    # ── Type Breakdown ─────────────────────────────────────────────────────────
    ws.cell(row=row_idx, column=1, value="TEST TYPE BREAKDOWN").font = Font(
        name="Segoe UI", bold=True, size=11, color="00F2FE")
    ws.cell(row=row_idx, column=1).fill = BG_HEADER
    ws.merge_cells(f"A{row_idx}:D{row_idx}")
    ws.cell(row=row_idx, column=1).alignment = ALIGN_CENTER
    ws.row_dimensions[row_idx].height = 20
    row_idx += 1
    for t_type, count in sorted(metrics["types"].items(), key=lambda x: -x[1]):
        pct = (count / metrics["total"] * 100) if metrics["total"] else 0
        color = TYPE_COLORS.get(t_type, "AAAAAA")
        for col_idx, val in enumerate([t_type, count, f"{pct:.1f}%", ""], start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = BG_ALT_ROW
            cell.font = Font(name="Segoe UI", size=9, color=color)
            cell.alignment = ALIGN_CENTER
            cell.border = thin_border()
        ws.row_dimensions[row_idx].height = 16
        row_idx += 1

    # Set remaining column widths for H-J
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 14

    return ws

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2 – DETAILS
# ─────────────────────────────────────────────────────────────────────────────

def build_details_sheet(wb, cases):
    ws = wb.create_sheet("Test Cases — Details")
    ws.sheet_view.showGridLines = False
    ws.tab_color = "B152FF"
    ws.freeze_panes = "A3"

    # Title
    ws.merge_cells("A1:K1")
    ws["A1"].value = "📋  SMART LEARN — SELENIUM TEST CASE DETAILS  (TC001 – TC305)"
    ws["A1"].font = Font(name="Segoe UI", bold=True, size=14, color="B152FF")
    ws["A1"].fill = BG_HEADER
    ws["A1"].alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 30

    # Column headers
    col_headers = [
        ("TC ID",        10),
        ("Module",       22),
        ("Feature",      20),
        ("Test Case Description", 48),
        ("Steps Summary",        38),
        ("Expected Result",      38),
        ("Priority",     10),
        ("Status",       10),
        ("Type",         14),
        ("Automation",   16),
        ("Remarks",      22),
    ]
    for col_idx, (hdr, width) in enumerate(col_headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=hdr)
        cell.font = FONT_HEADER
        cell.fill = PatternFill("solid", fgColor="1A1840")
        cell.alignment = ALIGN_CENTER
        cell.border = thin_border("4444AA")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 22

    # Data rows
    for row_idx, tc in enumerate(cases, start=3):
        tc_id, module, feature, desc, steps, expected, priority, status, tc_type, remarks = tc

        alt = (row_idx % 2 == 0)
        row_bg = MODULE_BG.get(module, BG_NORMAL if not alt else BG_ALT_ROW)

        status_font = FONT_PASS if status == "Pass" else (FONT_FAIL if status == "Fail" else FONT_SKIP)
        status_bg   = BG_PASS   if status == "Pass" else (BG_FAIL   if status == "Fail" else BG_SKIP)

        type_color = TYPE_COLORS.get(tc_type, "AAAAAA")

        row_data = [
            tc_id, module, feature, desc, steps, expected, priority, status, tc_type, "Selenium WebDriver 4", remarks
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border()

            if col_idx == 1:
                cell.font = FONT_ID
                cell.fill = BG_HEADER
                cell.alignment = ALIGN_CENTER
            elif col_idx == 2:
                cell.font = FONT_MODULE
                cell.fill = row_bg
                cell.alignment = ALIGN_LEFT
            elif col_idx == 3:
                cell.font = Font(name="Segoe UI", size=9, color="D0D0FF")
                cell.fill = row_bg
                cell.alignment = ALIGN_LEFT
            elif col_idx in (4, 5, 6):
                cell.font = FONT_BODY
                cell.fill = row_bg
                cell.alignment = ALIGN_LEFT
            elif col_idx == 7:
                cell.font = Font(name="Segoe UI", bold=True, size=9,
                                 color={"Critical":"FF4466","High":"FF8844","Medium":"FFCC44","Low":"88CC88"}.get(priority,"FFFFFF"))
                cell.fill = PRIORITY_COLORS.get(priority, BG_NORMAL)
                cell.alignment = ALIGN_CENTER
            elif col_idx == 8:
                cell.font = status_font
                cell.fill = status_bg
                cell.alignment = ALIGN_CENTER
            elif col_idx == 9:
                cell.font = Font(name="Segoe UI", bold=True, size=9, color=type_color)
                cell.fill = row_bg
                cell.alignment = ALIGN_CENTER
            elif col_idx == 10:
                cell.font = Font(name="Consolas", size=8, color="888899")
                cell.fill = BG_HEADER
                cell.alignment = ALIGN_CENTER
            else:
                cell.font = Font(name="Segoe UI", size=8, italic=True, color="666688")
                cell.fill = row_bg
                cell.alignment = ALIGN_LEFT

        ws.row_dimensions[row_idx].height = 42

    # Auto-filter
    ws.auto_filter.ref = f"A2:K{2 + len(cases)}"

    return ws

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3 – COVERAGE MATRIX
# ─────────────────────────────────────────────────────────────────────────────

def build_coverage_sheet(wb, metrics):
    ws = wb.create_sheet("Coverage Matrix")
    ws.sheet_view.showGridLines = False
    ws.tab_color = "F5820D"

    ws.merge_cells("A1:F1")
    ws["A1"].value = "🔬  MODULE × PRIORITY COVERAGE MATRIX"
    ws["A1"].font = Font(name="Segoe UI", bold=True, size=13, color="F5820D")
    ws["A1"].fill = BG_HEADER
    ws["A1"].alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 28

    priorities = ["Critical","High","Medium","Low"]
    headers = ["Module"] + priorities + ["Total"]
    for col_idx, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=hdr)
        cell.font = FONT_HEADER
        cell.fill = PatternFill("solid", fgColor="1A1840")
        cell.alignment = ALIGN_CENTER
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = 14 if col_idx > 1 else 26

    # Compute matrix
    matrix = {}
    for tc in TEST_CASES:
        m = tc[1]; p = tc[6]
        matrix.setdefault(m, {})
        matrix[m][p] = matrix[m].get(p, 0) + 1

    row_idx = 3
    for mod_name, mod_data in metrics["modules"].items():
        bg = MODULE_BG.get(mod_name, BG_NORMAL)
        ws.cell(row=row_idx, column=1, value=mod_name).font = FONT_MODULE
        ws.cell(row=row_idx, column=1).fill = bg
        ws.cell(row=row_idx, column=1).alignment = ALIGN_LEFT
        ws.cell(row=row_idx, column=1).border = thin_border()
        total_row = 0
        for col_p, p in enumerate(priorities, start=2):
            count = matrix.get(mod_name, {}).get(p, 0)
            total_row += count
            cell = ws.cell(row=row_idx, column=col_p, value=count)
            cell.fill = bg
            cell.font = Font(name="Segoe UI", bold=(count > 0), size=9,
                             color={"Critical":"FF4466","High":"FF8844","Medium":"FFCC44","Low":"88CC88"}[p] if count > 0 else "333355")
            cell.alignment = ALIGN_CENTER
            cell.border = thin_border()
        ws.cell(row=row_idx, column=6, value=total_row)
        ws.cell(row=row_idx, column=6).font = Font(name="Segoe UI", bold=True, size=9, color="00F2FE")
        ws.cell(row=row_idx, column=6).fill = bg
        ws.cell(row=row_idx, column=6).alignment = ALIGN_CENTER
        ws.cell(row=row_idx, column=6).border = thin_border()
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1

    # Totals row
    ws.cell(row=row_idx, column=1, value="TOTAL").font = Font(name="Segoe UI", bold=True, size=10, color="FFFFFF")
    ws.cell(row=row_idx, column=1).fill = PatternFill("solid", fgColor="1A1840")
    ws.cell(row=row_idx, column=1).alignment = ALIGN_CENTER
    ws.cell(row=row_idx, column=1).border = thin_border()
    for col_p, p in enumerate(priorities, start=2):
        total_p = sum(matrix.get(m, {}).get(p, 0) for m in matrix)
        cell = ws.cell(row=row_idx, column=col_p, value=total_p)
        cell.font = Font(name="Segoe UI", bold=True, size=10, color={"Critical":"FF4466","High":"FF8844","Medium":"FFCC44","Low":"88CC88"}[p])
        cell.fill = PatternFill("solid", fgColor="1A1840")
        cell.alignment = ALIGN_CENTER
        cell.border = thin_border()
    ws.cell(row=row_idx, column=6, value=sum(metrics["priorities"].values()))
    ws.cell(row=row_idx, column=6).font = Font(name="Segoe UI", bold=True, size=10, color="00F2FE")
    ws.cell(row=row_idx, column=6).fill = PatternFill("solid", fgColor="1A1840")
    ws.cell(row=row_idx, column=6).alignment = ALIGN_CENTER
    ws.cell(row=row_idx, column=6).border = thin_border()
    ws.row_dimensions[row_idx].height = 22

    return ws

# ─────────────────────────────────────────────────────────────────────────────
# Main Builder
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("🔧 Computing metrics…")
    metrics = compute_metrics(TEST_CASES)

    print(f"   ✅  Total Cases : {metrics['total']}")
    print(f"   ✅  Passed      : {metrics['passed']}")
    print(f"   ❌  Failed      : {metrics['failed']}")
    print(f"   ⏭  Skipped     : {metrics['skipped']}")
    print(f"   📊  Pass Rate   : {metrics['pass_rate']:.1f}%")

    print("📋 Building workbook…")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    build_summary_sheet(wb, metrics)
    build_details_sheet(wb, TEST_CASES)
    build_coverage_sheet(wb, metrics)

    output_file = "SmartLearn_Selenium_TestReport.xlsx"
    wb.save(output_file)
    print(f"✅  Workbook saved: {output_file}")
    print(f"   📄 Sheets: Summary | Test Cases — Details | Coverage Matrix")
    print(f"   📊 Total test cases documented: {metrics['total']} (TC001 – TC{metrics['total']:03d})")

if __name__ == "__main__":
    main()
